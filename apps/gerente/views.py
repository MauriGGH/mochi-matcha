"""
Panel de administración (rol gerente/admin) de Mochi Matcha.
Incluye: autenticación, floor plan, CRUD de menú/empleados/mesas,
reportes avanzados con exportación a Excel/PDF y gestión de imágenes.
"""
import json
import logging

logger = logging.getLogger(__name__)
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.cache import never_cache
from django.utils import timezone
from datetime import timedelta
from django.db import transaction
from django.db.models import Sum, Count
from django.db.models.functions import TruncDate, Coalesce
from django.db.models import DecimalField
from decimal import Decimal, InvalidOperation
from django.contrib import messages

from apps.accounts.decorators import gerente_requerido
from apps.accounts.models import Empleado
from apps.menu.models import Categoria, Producto, GrupoModificador, OpcionModificador, Promocion, TipoPromocion, TipoDescuento
from apps.mesas.models import Mesa, SesionCliente, UbicacionMesa
from apps.pedidos.models import Pedido, DetallePedido, SolicitudPago
from apps.auditoria.models import Auditoria
from apps.catalogs.models import MetodoPago, EstadoSolicitud
from apps.gerente.models import Configuracion


@never_cache
@ensure_csrf_cookie
def login_gerente(request):
    """Muestra y procesa el formulario de login exclusivo para gerente/admin.

    GET:  renderiza el formulario.
    POST: autentica al usuario; solo admite roles "gerente" y "admin".
    Redirige al dashboard si ya está autenticado con el rol correcto.
    """
    if request.user.is_authenticated and request.user.rol in ("gerente", "admin"):
        return redirect("gerente:dashboard")
    error = None
    if request.method == "POST":
        usuario = request.POST.get("usuario", "")
        contrasena = request.POST.get("contrasena", "")
        user = authenticate(request, usuario=usuario, password=contrasena)
        if user and user.rol in ("gerente", "admin") and user.is_active:
            login(request, user)
            return redirect("gerente:dashboard")
        error = "Credenciales incorrectas o sin acceso a este módulo."
    return render(request, "base/login.html", {
        "rol": "gerente", "rol_display": "Administrador",
        "form_action": "/gerente/login/", "error": error,
        "usuario_previo": request.POST.get("usuario", ""),
    })


def logout_gerente(request):
    """Cierra la sesión del gerente y redirige al login."""
    logout(request)
    return redirect("gerente:login_gerente")


# ─── Dashboard / Floor Plan ───────────────────────────────────────────────────

@gerente_requerido
@ensure_csrf_cookie
def dashboard(request):
    """Punto de entrada del panel; redirige directamente al floor plan."""
    return redirect("gerente:floor_plan")


@gerente_requerido
@ensure_csrf_cookie
def floor_plan(request):
    """Renderiza el plano del restaurante con el estado inicial de todas las mesas.

    Los contadores de listos y solicitudes se muestran en la barra superior
    y se actualizan en tiempo real por mesas_estado() vía polling JS.
    """
    mesas = Mesa.objects.prefetch_related("sesiones__pedidos").order_by("numero_mesa")
    listos_count = Pedido.objects.filter(estado="listo").count()
    solicitudes_count = SolicitudPago.objects.filter(
        estado_solicitud__descripcion="pendiente"
    ).count()
    return render(request, "gerente/floor_plan.html", {
        "mesas": mesas,
        "listos_count": listos_count,
        "solicitudes_count": solicitudes_count,
    })


@require_GET
@gerente_requerido
def stats_json(request):
    """Devuelve KPIs del día en JSON para el widget de estadísticas del dashboard.

    Retorna: ventas_hoy, pedidos_hoy, mesas_ocupadas, listos_count, solicitudes_count.
    """
    hoy = timezone.now().date()
    pedidos_hoy = Pedido.objects.filter(fecha_hora_ingreso__date=hoy).exclude(estado="cancelado")

    # BUG #4 FIX: reemplaza el doble bucle con Sum() agregado.
    # El código anterior hacía prefetch_related pero luego llamaba .all() dentro
    # del bucle interno, forzando queries adicionales por cada pedido (~N queries).
    # La agregación calcula el total en una sola query SQL.
    ventas_hoy = pedidos_hoy.aggregate(
        total=Sum("detalles__subtotal_calculado")
    )["total"] or 0.0

    return JsonResponse({
        "ventas_hoy": float(ventas_hoy),
        "pedidos_hoy": pedidos_hoy.count(),
        "mesas_ocupadas": Mesa.objects.filter(estado="ocupada").count(),
        "listos_count": Pedido.objects.filter(estado="listo").count(),
        "solicitudes_count": SolicitudPago.objects.filter(
            estado_solicitud__descripcion="pendiente"
        ).count(),
    })


# ─── Detalle de mesa ─────────────────────────────────────────────────────────

@require_GET
@gerente_requerido
def detalle_mesa(request, mesa_id):
    """Devuelve JSON con el detalle completo de una mesa: sesiones activas, pedidos y solicitudes.

    Parámetros URL:
        mesa_id: PK de la Mesa.
    Retorna JSON con: ok, mesa_id, numero_mesa, pin, estado, mesa_libre,
        sesiones (lista con pedidos e ítems), solicitudes pendientes y total_mesa.
    """
    mesa = get_object_or_404(Mesa, pk=mesa_id)
    sesiones = mesa.sesiones.filter(estado="activa").order_by("fecha_inicio")
    sesiones_data = []
    for s in sesiones:
        pedidos_sesion = s.pedidos.exclude(estado="cancelado").prefetch_related(
            "detalles__producto", "detalles__modificadores__opcion"
        ).order_by("-fecha_hora_ingreso")
        total_sesion = sum(
            sum(d.subtotal_calculado for d in p.detalles.all()) for p in pedidos_sesion
        )
        sesiones_data.append({
            "id": s.pk,
            "alias": s.alias,
            "total": float(total_sesion),
            "pedidos": [
                {
                    "id": p.pk,
                    "estado": p.estado,
                    "estado_display": p.get_estado_display(),
                    "fecha": p.fecha_hora_ingreso.strftime("%H:%M"),
                    "items": [
                        {
                            "nombre": d.producto.nombre,
                            "cantidad": d.cantidad,
                            "subtotal": float(d.subtotal_calculado),
                            "notas": d.notas or "",
                            "modificadores": [m.opcion.nombre_opcion for m in d.modificadores.all()],
                        }
                        for d in p.detalles.all()
                    ],
                }
                for p in pedidos_sesion
            ],
        })

    solicitudes = []
    for s in sesiones:
        for sol in s.solicitudes_pago.filter(estado_solicitud__descripcion="pendiente"):
            solicitudes.append({
                "id": sol.pk,
                "alias": s.alias,
                "tipo": sol.tipo,
                "tipo_display": sol.get_tipo_display(),
                "total": float(sol.total_mesa or sol.total_individual or 0),
                "fecha": sol.fecha_hora.strftime("%H:%M"),
                "sesion_id": s.pk,
                "metodo_pref": sol.detalle_pago or "",
            })

    total_mesa = sum(s["total"] for s in sesiones_data)
    return JsonResponse({
        "ok": True,
        "mesa_id": mesa.pk,
        "numero_mesa": mesa.numero_mesa,
        "pin": mesa.pin_actual or "",
        "estado": mesa.estado,
        "mesa_libre": mesa.estado == "libre",
        "sesiones": sesiones_data,
        "solicitudes": solicitudes,
        "total_mesa": total_mesa,
    })


# ─── Cancelar pedido (solo gerente) ──────────────────────────────────────────

@require_POST
@gerente_requerido
def cancelar_pedido(request):
    """POST JSON: { pedido_id, motivo }"""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "JSON inválido"}, status=400)

    pedido_id = data.get("pedido_id")
    motivo = data.get("motivo", "").strip()

    if not motivo:
        return JsonResponse({"ok": False, "error": "El motivo es obligatorio"}, status=400)

    pedido = get_object_or_404(Pedido, pk=pedido_id)
    if pedido.estado in ("entregado", "cancelado"):
        return JsonResponse({"ok": False, "error": "No se puede cancelar este pedido"}, status=400)

    pedido.estado = "cancelado"
    pedido.motivo_cancelacion = motivo
    pedido.save(update_fields=["estado", "motivo_cancelacion"])

    Auditoria.objects.create(
        accion="Pedido cancelado",
        detalle=f"Pedido #{pedido.pk} cancelado. Motivo: {motivo}",
        empleado=request.user,
        mesa=pedido.sesion.mesa,
        pedido=pedido,
    )

    return JsonResponse({"ok": True})


# ─── Cancelar solicitud de pago (gerente) ─────────────────────────────────────

@require_POST
@gerente_requerido
def cancelar_solicitud_pago(request):
    """
    POST JSON {solicitud_id}
    Cancela una SolicitudPago pendiente. Usa select_for_update para evitar
    condiciones de carrera.
    """
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "JSON inválido"}, status=400)

    solicitud_id = data.get("solicitud_id")
    if not solicitud_id:
        return JsonResponse({"ok": False, "error": "solicitud_id requerido"}, status=400)

    with transaction.atomic():
        try:
            solicitud = (
                SolicitudPago.objects
                .select_for_update(nowait=True)
                .select_related("estado_solicitud", "mesa", "sesion")
                .get(pk=solicitud_id)
            )
        except SolicitudPago.DoesNotExist:
            return JsonResponse({"ok": False, "error": "Solicitud no encontrada"}, status=404)
        except Exception:
            return JsonResponse({"ok": False, "error": "La solicitud está siendo procesada. Intenta de nuevo."}, status=409)

        if solicitud.estado_solicitud.descripcion != "pendiente":
            return JsonResponse({
                "ok": False,
                "error": f"No se puede cancelar: la solicitud está en estado '{solicitud.estado_solicitud.descripcion}'."
            }, status=400)

        estado_cancelada, _ = EstadoSolicitud.objects.get_or_create(descripcion="cancelada")
        solicitud.estado_solicitud = estado_cancelada
        solicitud.save(update_fields=["estado_solicitud"])

        Auditoria.objects.create(
            accion="Solicitud de pago cancelada (gerente)",
            detalle=(
                f"Solicitud #{solicitud.pk} (mesa {solicitud.mesa.numero_mesa if solicitud.mesa else 'N/A'}) "
                f"cancelada por gerente. Tipo: {solicitud.tipo}. "
                f"Total: ${solicitud.total_mesa or solicitud.total_individual or 0:.2f}"
            ),
            empleado=request.user,
            mesa=solicitud.mesa,
            solicitud_pago=solicitud,
        )

    return JsonResponse({"ok": True, "mensaje": "Solicitud cancelada correctamente."})


# ─── Mesas ────────────────────────────────────────────────────────────────────

@gerente_requerido
def mesas_estado(request):
    """JSON polling para el floor plan del gerente."""
    # BUG #5 FIX: prefetch completo para evitar N+1 queries por polling.
    # Igual que en mesero/views.py, las llamadas anidadas a count()/exists()
    # generaban ~100 queries por cada actualización del floor plan.
    mesas = Mesa.objects.prefetch_related(
        "sesiones__pedidos",
        "sesiones__solicitudes_pago__estado_solicitud",
        "alertas",
    ).order_by("numero_mesa")

    from apps.mesas.models import AlertaMesero
    listos_count = Pedido.objects.filter(estado="listo").count()
    solicitudes_count = SolicitudPago.objects.filter(
        estado_solicitud__descripcion="pendiente"
    ).count()
    alertas_count = AlertaMesero.objects.filter(atendida=False).count()

    data = []
    for m in mesas:
        sesiones_activas = [s for s in m.sesiones.all() if s.estado == "activa"]
        pedidos_cocina = 0
        pedidos_listos = 0
        tiene_solicitud = False
        tiene_alerta = any(not a.atendida for a in m.alertas.all())

        for s in sesiones_activas:
            for p in s.pedidos.all():
                if p.estado in ("recibido", "preparando"):
                    pedidos_cocina += 1
                elif p.estado == "listo":
                    pedidos_listos += 1
            for sol in s.solicitudes_pago.all():
                if sol.estado_solicitud.descripcion == "pendiente":
                    tiene_solicitud = True
                    break

        if m.estado == "libre":
            estado_visual = "libre"
        elif tiene_alerta:
            estado_visual = "alerta"
        elif tiene_solicitud:
            estado_visual = "cobrando"
        elif pedidos_listos > 0:
            estado_visual = "listo"
        elif pedidos_cocina > 0:
            estado_visual = "cocina"
        else:
            estado_visual = "ocupada"

        data.append({
            "id": m.pk,
            "numero": m.numero_mesa,
            "estado": m.estado,
            "estado_visual": estado_visual,
            "pin": m.pin_actual or "",
            "clientes": len(sesiones_activas),
            "pedidos_cocina": pedidos_cocina,
            "pedidos_listos": pedidos_listos,
            "tiene_solicitud": tiene_solicitud,
            "tiene_alerta": tiene_alerta,
        })
    return JsonResponse({
        "ok": True, "mesas": data,
        "listos_count": listos_count,
        "solicitudes_count": solicitudes_count,
        "alertas_count": alertas_count,
    })


@gerente_requerido
def mesas(request):
    """Alias de compatibilidad; redirige al CRUD de mesas."""
    return redirect("gerente:mesas_crud")


@gerente_requerido
def mesas_crud(request):
    """Lista todas las mesas con su QR y permite crear nuevas.

    POST: crea una mesa con número único y QR generado automáticamente.
    GET:  renderiza la vista de gestión con QR en base64 para cada mesa.
    """
    if request.method == "POST":
        numero = request.POST.get("numero_mesa")
        capacidad = request.POST.get("capacidad", 4)
        ubicacion_id = request.POST.get("ubicacion_id") or None
        import uuid as _uuid
        qr = f"mesa-{numero}-{_uuid.uuid4().hex[:8]}"
        ubicacion_obj = UbicacionMesa.objects.filter(pk=ubicacion_id).first() if ubicacion_id else None
        Mesa.objects.get_or_create(
            numero_mesa=numero,
            defaults={"capacidad": capacidad, "ubicacion": ubicacion_obj, "codigo_qr": qr}
        )
        return redirect("gerente:mesas_crud")

    all_mesas = Mesa.objects.select_related("ubicacion").order_by("numero_mesa")
    # QR generado con el dominio REAL del request — funciona en producción al
    # cambiar de dominio sin tener que reconfigurar SITE_BASE_URL.
    qr_base_url = request.build_absolute_uri("/").rstrip("/")
    for mesa in all_mesas:
        mesa.qr_img = mesa.generate_qr_base64(qr_base_url)
    meseros = Empleado.objects.filter(rol="mesero", activo=True)
    ubicaciones = UbicacionMesa.objects.order_by("nombre")
    return render(request, "gerente/menu_manager.html", {
        "mesas": all_mesas, "meseros": meseros,
        "ubicaciones": ubicaciones, "vista": "mesas"
    })


@require_POST
@gerente_requerido
def mesa_editar(request, id):
    """Edita número, capacidad y ubicación de una mesa existente."""
    mesa = get_object_or_404(Mesa, pk=id)
    numero       = request.POST.get("numero_mesa")
    capacidad    = request.POST.get("capacidad")
    ubicacion_id = request.POST.get("ubicacion_id") or None

    if numero:
        try:
            numero_int = int(numero)
        except (TypeError, ValueError):
            messages.error(request, "El número de mesa no es válido.")
            return redirect("gerente:mesas_crud")
        # El número de mesa es único: bloquear colisiones con otra mesa.
        if Mesa.objects.exclude(pk=mesa.pk).filter(numero_mesa=numero_int).exists():
            messages.error(request, f"Ya existe otra mesa con el número {numero_int}.")
            return redirect("gerente:mesas_crud")
        mesa.numero_mesa = numero_int

    if capacidad:
        try:
            mesa.capacidad = max(1, int(capacidad))
        except (TypeError, ValueError):
            messages.error(request, "La capacidad no es válida.")
            return redirect("gerente:mesas_crud")

    mesa.ubicacion = UbicacionMesa.objects.filter(pk=ubicacion_id).first() if ubicacion_id else None
    mesa.save(update_fields=["numero_mesa", "capacidad", "ubicacion"])
    messages.success(request, f"Mesa {mesa.numero_mesa} actualizada.")
    return redirect("gerente:mesas_crud")


@require_POST
@gerente_requerido
def ubicacion_crear(request):
    """AJAX: crea una UbicacionMesa y devuelve JSON {ok, id, nombre}."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "JSON inválido"}, status=400)
    nombre = data.get("nombre", "").strip()
    if not nombre:
        return JsonResponse({"ok": False, "error": "El nombre es obligatorio"}, status=400)
    obj, created = UbicacionMesa.objects.get_or_create(nombre=nombre)
    return JsonResponse({"ok": True, "id": obj.pk, "nombre": obj.nombre, "created": created})


@require_POST
@gerente_requerido
def ubicacion_editar(request, id):
    """AJAX: edita el nombre de una UbicacionMesa."""
    obj = get_object_or_404(UbicacionMesa, pk=id)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "JSON inválido"}, status=400)
    nombre = data.get("nombre", "").strip()
    if not nombre:
        return JsonResponse({"ok": False, "error": "El nombre es obligatorio"}, status=400)
    obj.nombre = nombre  # save() del modelo aplica .upper()
    obj.save()
    return JsonResponse({"ok": True, "id": obj.pk, "nombre": obj.nombre})


@require_POST
@gerente_requerido
def ubicacion_eliminar(request, id):
    """AJAX: elimina una UbicacionMesa solo si no tiene mesas asociadas."""
    obj = get_object_or_404(UbicacionMesa, pk=id)
    if obj.mesas.exists():
        return JsonResponse(
            {"ok": False, "error": "No se puede eliminar: tiene mesas asociadas."},
            status=400
        )
    obj.delete()
    return JsonResponse({"ok": True})


@require_POST
@gerente_requerido
def mesa_eliminar(request, id):
    # P8: bloquear la mesa y re-validar su estado dentro de la transacción —
    # evita borrar una mesa que justo pasó a "ocupada" en otra request.
    with transaction.atomic():
        mesa = get_object_or_404(Mesa.objects.select_for_update(), pk=id)
        if mesa.estado == "libre" and not mesa.sesiones.filter(estado="activa").exists():
            mesa.delete()
        else:
            messages.error(request, f"No se puede eliminar la Mesa {mesa.numero_mesa}: está ocupada.")
    return redirect("gerente:mesas_crud")


@require_POST
@gerente_requerido
def asignar_mesero(request, mesa_id):
    """Asigna o desasigna un mesero a una mesa.

    POST form: mesero_id (vacío para desasignar).
    Solo acepta empleados con rol "mesero".
    """
    mesa = get_object_or_404(Mesa, pk=mesa_id)
    mesero_id = request.POST.get("mesero_id")
    if mesero_id:
        mesero = get_object_or_404(Empleado, pk=mesero_id, rol="mesero")
        mesa.id_mesero_asignado = mesero
    else:
        mesa.id_mesero_asignado = None
    mesa.save(update_fields=["id_mesero_asignado"])
    return JsonResponse({"ok": True})


# ─── Productos ────────────────────────────────────────────────────────────────

@gerente_requerido
def productos(request):
    """Lista todos los productos ordenados por categoría y nombre."""
    prods = Producto.objects.select_related("categoria").order_by("categoria__orden", "nombre")
    categorias = Categoria.objects.order_by("orden")
    return render(request, "gerente/menu_manager.html", {
        "productos": prods, "categorias": categorias, "vista": "productos"
    })


@gerente_requerido
def productos_nuevo(request):
    """Crea un nuevo producto en el menú.

    GET:  muestra el formulario de alta.
    POST: valida y persiste el nuevo producto.
    """
    categorias = Categoria.objects.order_by("orden")
    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        precio = request.POST.get("precio", 0)
        descripcion = request.POST.get("descripcion", "").strip()
        imagen_url = request.POST.get("imagen_url", "").strip()
        categoria_id = request.POST.get("categoria_id")
        disponible = request.POST.get("disponible") == "on"
        cat = get_object_or_404(Categoria, pk=categoria_id)
        Producto.objects.create(
            nombre=nombre, precio=precio, descripcion=descripcion or None,
            imagen_url=imagen_url or None, categoria=cat, disponible=disponible,
        )
        return redirect("gerente:productos")
    return render(request, "gerente/menu_manager.html", {
        "categorias": categorias, "vista": "productos", "form_nuevo": True
    })


@gerente_requerido
def producto_editar(request, id):
    """Edita un producto existente.

    GET:  precarga los datos del producto en el formulario.
    POST: actualiza nombre, precio, descripción, imagen, disponibilidad y categoría.
    """
    producto = get_object_or_404(Producto, pk=id)
    categorias = Categoria.objects.order_by("orden")
    if request.method == "POST":
        producto.nombre = request.POST.get("nombre", producto.nombre).strip()
        producto.precio = request.POST.get("precio", producto.precio)
        producto.descripcion = request.POST.get("descripcion", "").strip() or None
        producto.imagen_url = request.POST.get("imagen_url", "").strip() or None
        producto.disponible = request.POST.get("disponible") == "on"
        cat_id = request.POST.get("categoria_id")
        if cat_id:
            producto.categoria = get_object_or_404(Categoria, pk=cat_id)
        producto.save()
        return redirect("gerente:productos")
    return render(request, "gerente/menu_manager.html", {
        "producto_editar": producto, "categorias": categorias, "vista": "productos"
    })


@require_POST
@gerente_requerido
def producto_eliminar(request, id):
    """Soft-delete: marca el producto como no disponible en lugar de borrarlo físicamente.

    Preserva la integridad referencial con DetallePedido histórico.
    """
    producto = get_object_or_404(Producto, pk=id)
    producto.disponible = False
    producto.save(update_fields=["disponible"])
    return JsonResponse({"ok": True})


# ─── Categorías ───────────────────────────────────────────────────────────────

@gerente_requerido
def categorias(request):
    """Lista categorías y permite crear nuevas.

    GET:  renderiza la lista de categorías.
    POST: crea una nueva categoría con nombre, orden y área (cocina/barra/ambos).
    """
    cats = Categoria.objects.order_by("orden")
    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        orden = int(request.POST.get("orden", 0))
        area = request.POST.get("area", "ambos")
        if nombre:
            Categoria.objects.create(nombre=nombre, orden=orden, area=area)
        return redirect("gerente:categorias")
    return render(request, "gerente/menu_manager.html", {
        "categorias": cats, "vista": "categorias"
    })


@require_POST
@gerente_requerido
def categoria_editar(request, id):
    """Edita nombre, orden y área de una categoría existente."""
    cat = get_object_or_404(Categoria, pk=id)
    nombre = request.POST.get("nombre", "").strip()
    if not nombre:
        return JsonResponse({"ok": False, "error": "El nombre es obligatorio"}, status=400)

    try:
        orden = int(request.POST.get("orden", cat.orden))
    except (TypeError, ValueError):
        orden = cat.orden

    area = request.POST.get("area", cat.area)
    if area not in dict(Categoria.AREAS):
        area = cat.area

    cat.nombre = nombre   # save() del modelo aplica .upper()
    cat.orden = orden
    cat.area = area
    cat.save()
    return JsonResponse({
        "ok": True, "id": cat.pk, "nombre": cat.nombre,
        "orden": cat.orden, "area": cat.area,
    })


@require_POST
@gerente_requerido
def categoria_eliminar(request, id):
    """Elimina una categoría solo si no tiene productos activos asociados.

    Evita dejar productos huérfanos sin categoría visible en el menú.
    """
    cat = get_object_or_404(Categoria, pk=id)
    if not cat.productos.filter(disponible=True).exists():
        cat.delete()
        return JsonResponse({"ok": True})
    return JsonResponse({"ok": False, "error": "Tiene productos activos asociados"}, status=400)


# ─── Modificadores ────────────────────────────────────────────────────────────

@gerente_requerido
def modificadores(request):
    """Lista todos los grupos de modificadores y las plantillas disponibles para clonar."""
    grupos = GrupoModificador.objects.prefetch_related("productos", "opciones").order_by("-pk")
    productos_list = Producto.objects.filter(disponible=True).order_by("nombre")
    plantillas = GrupoModificador.objects.filter(es_plantilla=True).prefetch_related("opciones").order_by("-pk")
    return render(request, "gerente/menu_manager.html", {
        "grupos": grupos, "productos_list": productos_list,
        "plantillas": plantillas, "vista": "modificadores"
    })


@require_POST
@gerente_requerido
def modificador_crear(request):
    """Crea un GrupoModificador con sus opciones y lo asocia a uno o más productos.

    POST JSON: { nombre_grupo, tipo, es_obligatorio, max_selecciones,
                 producto_ids[], opciones[{ nombre, precio_extra }] }
    Retorna JSON {ok, grupo_id}.
    """
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"ok": False}, status=400)

    opciones = data.get("opciones", [])
    if not opciones:
        return JsonResponse(
            {"ok": False, "error": "Debe agregar al menos una opción al grupo"},
            status=400,
        )

    # Soporta producto_ids (lista M2M) o producto_id (único, retrocompatible)
    producto_ids = data.get("producto_ids") or []
    if not producto_ids:
        single = data.get("producto_id")
        if single:
            producto_ids = [single]

    if not producto_ids:
        return JsonResponse({"ok": False, "error": "Selecciona al menos un producto"}, status=400)

    productos = list(Producto.objects.filter(pk__in=producto_ids, disponible=True))
    if not productos:
        return JsonResponse({"ok": False, "error": "Productos no encontrados"}, status=400)

    grupo = GrupoModificador.objects.create(
        nombre_grupo=data.get("nombre_grupo", "").strip(),
        tipo=data.get("tipo", "única"),
        es_obligatorio=data.get("es_obligatorio", False),
        max_selecciones=data.get("max_selecciones") or None,
    )
    grupo.productos.set(productos)

    for op in opciones:
        OpcionModificador.objects.create(
            nombre_opcion=op.get("nombre", "").strip(),
            precio_extra=op.get("precio_extra", 0),
            grupo=grupo,
        )
    return JsonResponse({"ok": True, "grupo_id": grupo.pk})


@require_POST
@gerente_requerido
def modificador_eliminar(request, id):
    """Elimina un GrupoModificador y sus opciones si no están protegidas."""
    grupo = get_object_or_404(GrupoModificador, pk=id)
    grupo.delete()
    return JsonResponse({"ok": True})


@require_POST
@gerente_requerido
def modificador_toggle_plantilla(request, id):
    """Marca/desmarca un grupo como plantilla reutilizable."""
    grupo = get_object_or_404(GrupoModificador, pk=id)
    grupo.es_plantilla = not grupo.es_plantilla
    grupo.save(update_fields=["es_plantilla"])
    return JsonResponse({"ok": True, "es_plantilla": grupo.es_plantilla})


@require_POST
@gerente_requerido
def modificador_clonar(request):
    """Clona un grupo plantilla añadiéndolo a uno o más productos."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "JSON inválido"}, status=400)
    plantilla_id = data.get("plantilla_id")
    producto_ids = data.get("producto_ids") or []
    if not producto_ids:
        single = data.get("producto_id")
        if single:
            producto_ids = [single]
    plantilla = get_object_or_404(GrupoModificador, pk=plantilla_id, es_plantilla=True)
    productos = list(Producto.objects.filter(pk__in=producto_ids, disponible=True))
    if not productos:
        return JsonResponse({"ok": False, "error": "Productos no encontrados"}, status=400)
    nuevo = GrupoModificador.objects.create(
        nombre_grupo=plantilla.nombre_grupo,
        tipo=plantilla.tipo,
        es_obligatorio=plantilla.es_obligatorio,
        max_selecciones=plantilla.max_selecciones,
        es_plantilla=False,
    )
    nuevo.productos.set(productos)
    for op in plantilla.opciones.all():
        op.__class__.objects.create(
            nombre_opcion=op.nombre_opcion,
            precio_extra=op.precio_extra,
            grupo=nuevo,
        )
    return JsonResponse({"ok": True, "grupo_id": nuevo.pk})


@gerente_requerido
def modificador_editar(request, id):
    """
    GET  → devuelve JSON con datos completos del grupo (nombre, tipo, productos, opciones).
    POST → actualiza el grupo usando estrategia upsert segura para opciones:
           - Opciones con 'id' en el payload → se actualizan.
           - Opciones sin 'id' → se crean nuevas.
           - Opciones que ya no están en el payload → se intentan eliminar;
             si tienen usos históricos (ProtectedError) se conservan intactas
             para no romper la integridad referencial de DetalleModificador.
    Justificación: DetalleModificador.opcion usa on_delete=PROTECT; borrar
    opciones usadas lanzaría ProtectedError. Las opciones obsoletas sin usos
    sí se eliminan para mantener el catálogo limpio.
    """
    from django.db import IntegrityError
    from django.db.models import ProtectedError

    grupo = get_object_or_404(GrupoModificador, pk=id)

    if request.method == "GET":
        # Solo opciones activas para el modal de edición; las inactivas son histórico
        opciones = [
            {"id": op.pk, "nombre": op.nombre_opcion, "precio_extra": float(op.precio_extra)}
            for op in grupo.opciones.filter(activo=True)
        ]
        producto_ids = list(grupo.productos.values_list("pk", flat=True))
        return JsonResponse({
            "ok": True,
            "grupo": {
                "id":              grupo.pk,
                "nombre_grupo":    grupo.nombre_grupo,
                "tipo":            grupo.tipo,
                "es_obligatorio":  grupo.es_obligatorio,
                "max_selecciones": grupo.max_selecciones,
                "producto_ids":    producto_ids,
                "opciones":        opciones,
            }
        })

    # ── POST: actualizar ─────────────────────────────────────────────────────
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "JSON inválido"}, status=400)

    opciones_payload = data.get("opciones", [])
    if not opciones_payload:
        return JsonResponse({"ok": False, "error": "Debe agregar al menos una opción"}, status=400)

    producto_ids = data.get("producto_ids") or []
    if not producto_ids:
        single = data.get("producto_id")
        if single:
            producto_ids = [single]
    if not producto_ids:
        return JsonResponse({"ok": False, "error": "Selecciona al menos un producto"}, status=400)

    productos = list(Producto.objects.filter(pk__in=producto_ids, disponible=True))
    if not productos:
        return JsonResponse({"ok": False, "error": "Productos no encontrados"}, status=400)

    with transaction.atomic():
        # 1. Actualizar campos del grupo
        grupo.nombre_grupo    = data.get("nombre_grupo", grupo.nombre_grupo).strip()
        grupo.tipo            = data.get("tipo", grupo.tipo)
        grupo.es_obligatorio  = data.get("es_obligatorio", grupo.es_obligatorio)
        grupo.max_selecciones = data.get("max_selecciones") or None
        grupo.save()

        # 2. Actualizar productos (M2M)
        grupo.productos.set(productos)

        # 3. Upsert de opciones (estrategia segura)
        ids_en_payload = set()
        for op_data in opciones_payload:
            nombre = op_data.get("nombre", "").strip()
            precio = op_data.get("precio_extra", 0)
            op_id  = op_data.get("id")

            if op_id:
                # Actualizar existente si pertenece al grupo
                OpcionModificador.objects.filter(pk=op_id, grupo=grupo).update(
                    nombre_opcion=nombre,
                    precio_extra=precio,
                )
                ids_en_payload.add(int(op_id))
            else:
                # Crear nueva opción
                nueva = OpcionModificador.objects.create(
                    nombre_opcion=nombre,
                    precio_extra=precio,
                    grupo=grupo,
                )
                ids_en_payload.add(nueva.pk)

        # 4. Opciones que ya no están en el payload:
        #    - Si no tienen usos → eliminar físicamente (limpia el catálogo)
        #    - Si tienen usos (ProtectedError) → soft-delete: activo=False
        opciones_a_eliminar = grupo.opciones.exclude(pk__in=ids_en_payload)
        for op in opciones_a_eliminar:
            try:
                op.delete()
            except ProtectedError:
                # Tiene usos en DetalleModificador → marcar inactiva para preservar histórico
                op.activo = False
                op.save(update_fields=["activo"])

    return JsonResponse({"ok": True, "grupo_id": grupo.pk})


# ─── Promociones ──────────────────────────────────────────────────────────────

def _parse_promo_form(post):
    """
    2.1: valida y normaliza el POST del modal de promoción.

    Devuelve (datos, errores). `datos` contiene los valores ya casteados al tipo
    correcto cuando son válidos. `errores` es una lista de mensajes en español
    apta para mostrar al gerente vía `messages.error`. El llamador decide si
    aplicar los cambios.

    Reglas (espejo del modelo Promocion + del validador client-side):
      - titulo: 3-100 chars (después de strip).
      - descripcion_corta: hasta 120 chars o vacío.
      - orden: entero 0-32767.
      - tipo_descuento_id: requerido, debe existir.
      - valor_descuento: requerido salvo "2x1"; Decimal con ≤ 2 decimales;
        "Porcentaje" → 1-100; resto → 0-99999999.99.
      - cantidad_minima: requerido para "Lleva X paga Y"; entero ≥ 2 y > Y.
      - fecha_inicio / fecha_fin: requeridos; fin > inicio.
      - aplicacion: una de las choices del modelo.
      - imagen_url: hasta 500 chars.
      - dias_semana: solo "0".."6", CSV ordenado.
      - requiere_todos_productos: requiere ≥ 2 productos_aplicables.
    """
    errs = []

    titulo = (post.get("titulo") or "").strip()
    if len(titulo) < 3:
        errs.append("El título debe tener al menos 3 caracteres.")
    if len(titulo) > 100:
        errs.append("El título no puede superar 100 caracteres.")

    descripcion_corta = (post.get("descripcion_corta") or "").strip() or None
    if descripcion_corta and len(descripcion_corta) > 120:
        errs.append("La descripción corta no puede superar 120 caracteres.")

    try:
        orden = int(post.get("orden", "0") or "0")
        if orden < 0 or orden > 32767:
            errs.append("El orden debe estar entre 0 y 32767.")
            orden = 0
    except (ValueError, TypeError):
        errs.append("El orden debe ser un número entero.")
        orden = 0

    tipo_descuento_id = post.get("tipo_descuento_id")
    tipo = None
    tipo_desc = ""
    if not tipo_descuento_id:
        errs.append("Selecciona un tipo de descuento.")
    else:
        tipo = TipoDescuento.objects.filter(pk=tipo_descuento_id).first()
        if not tipo:
            errs.append("Tipo de descuento inválido.")
        else:
            tipo_desc = tipo.descripcion

    aplicacion = post.get("aplicacion", "item")
    if aplicacion not in {"item", "total", "combo"}:
        errs.append("Aplicación inválida.")
        aplicacion = "item"

    # valor_descuento: requerido salvo "2x1".
    valor_raw = (post.get("valor_descuento") or "").strip()
    valor_decimal = None
    if tipo_desc != "2x1":
        if not valor_raw:
            errs.append("Ingresa el valor del descuento.")
        else:
            try:
                valor_decimal = Decimal(valor_raw)
            except (InvalidOperation, ValueError):
                errs.append("El valor del descuento debe ser un número.")
            else:
                if valor_decimal < 0:
                    errs.append("El valor del descuento no puede ser negativo.")
                if valor_decimal.as_tuple().exponent < -2:
                    errs.append("El valor del descuento admite máximo 2 decimales.")
                if valor_decimal > Decimal("99999999.99"):
                    errs.append("El valor del descuento excede el máximo permitido.")
                if tipo_desc == "Porcentaje" and not (Decimal("0") < valor_decimal <= Decimal("100")):
                    errs.append("El porcentaje debe estar entre 1 y 100.")

    # cantidad_minima: requerido para "Lleva X paga Y".
    cant_raw = (post.get("cantidad_minima") or "").strip()
    cantidad_int = None
    if cant_raw:
        try:
            cantidad_int = int(cant_raw)
            if cantidad_int < 0:
                errs.append("La cantidad mínima no puede ser negativa.")
        except (ValueError, TypeError):
            errs.append("La cantidad mínima debe ser un entero.")
    if tipo_desc == "Lleva X paga Y":
        if cantidad_int is None or cantidad_int < 2:
            errs.append("Para 'Lleva X paga Y' la cantidad llevada (X) debe ser ≥ 2.")
        if (
            cantidad_int is not None and valor_decimal is not None
            and cantidad_int <= valor_decimal
        ):
            errs.append("La cantidad llevada (X) debe ser mayor que la pagada (Y).")

    fecha_inicio = post.get("fecha_inicio")
    fecha_fin    = post.get("fecha_fin")
    if not fecha_inicio:
        errs.append("Indica la fecha de inicio.")
    if not fecha_fin:
        errs.append("Indica la fecha de fin.")
    if fecha_inicio and fecha_fin and fecha_inicio >= fecha_fin:
        errs.append("La fecha fin debe ser posterior al inicio.")

    imagen_url = (post.get("imagen_url") or "").strip() or None
    if imagen_url and len(imagen_url) > 500:
        errs.append("La URL de la imagen no puede superar 500 caracteres.")

    dias_validos = sorted({
        d for d in post.getlist("dias_semana")
        if d in {"0", "1", "2", "3", "4", "5", "6"}
    })
    dias_csv = ",".join(dias_validos)

    requiere_todos = post.get("requiere_todos_productos") == "1"
    ids_aplicables = post.getlist("productos_aplicables")
    if requiere_todos and len(ids_aplicables) < 2:
        errs.append("El modo combo requiere al menos 2 productos aplicables.")

    datos = {
        "titulo": titulo,
        "descripcion_corta": descripcion_corta,
        "orden": orden,
        "tipo": tipo,
        "aplicacion": aplicacion,
        "valor_decimal": valor_decimal,
        "cantidad_int": cantidad_int,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "imagen_url": imagen_url,
        "dias_csv": dias_csv,
        "requiere_todos": requiere_todos,
        "ids_aplicables": ids_aplicables,
        "ids_beneficiados": post.getlist("productos_beneficiados"),
    }
    return datos, errs


@gerente_requerido
def promociones(request):
    """Lista promociones y permite crear nuevas.

    POST: crea una promoción con tipo de descuento, rango de fechas, días de la
          semana habilitados y productos aplicables/beneficiados (M2M).
    GET:  renderiza la lista con los datos necesarios para los formularios del modal.
    """
    if request.method == "POST":
        datos, errs = _parse_promo_form(request.POST)
        if errs:
            for e in errs:
                messages.error(request, e)
            return redirect("gerente:promociones")

        promocion = Promocion.objects.create(
            titulo=datos["titulo"],
            tipo_descuento=datos["tipo"],
            valor_descuento=datos["valor_decimal"],
            cantidad_minima=datos["cantidad_int"],
            aplicacion=datos["aplicacion"],
            fecha_inicio=datos["fecha_inicio"],
            fecha_fin=datos["fecha_fin"],
            activa=True,
            imagen_url=datos["imagen_url"],
            descripcion_corta=datos["descripcion_corta"],
            orden=datos["orden"],
            dias_semana=datos["dias_csv"],
            requiere_todos_productos=datos["requiere_todos"],
        )

        if datos["ids_aplicables"]:
            promocion.productos_aplicables.set(datos["ids_aplicables"])
        if datos["ids_beneficiados"]:
            promocion.productos_beneficiados.set(datos["ids_beneficiados"])

        messages.success(request, f"Promoción '{datos['titulo']}' creada exitosamente.")
        return redirect("gerente:promociones")

    # GET
    promos = Promocion.objects.select_related("tipo_descuento").prefetch_related(
        "productos_aplicables"
    ).order_by("-fecha_inicio")
    tipos_descuento = TipoDescuento.objects.all()
    productos_list = Producto.objects.filter(disponible=True).order_by("nombre")
    return render(request, "gerente/menu_manager.html", {
        "promociones": promos,
        "tipos_descuento": tipos_descuento,
        "productos_list": productos_list,
        "vista": "promociones",
    })


@require_POST
@gerente_requerido
def promocion_toggle(request, id):
    """Activa o desactiva una promoción sin borrarla. Retorna JSON {ok, activa}."""
    promo = get_object_or_404(Promocion, pk=id)
    promo.activa = not promo.activa
    promo.save(update_fields=["activa"])
    return JsonResponse({"ok": True, "activa": promo.activa})


@require_POST
@gerente_requerido
def promocion_eliminar(request, id):
    """Elimina permanentemente una promoción."""
    promo = get_object_or_404(Promocion, pk=id)
    promo.delete()
    return JsonResponse({"ok": True})


@gerente_requerido
def promocion_editar(request, id):
    """Edita una promoción existente.

    GET:  devuelve JSON con todos los datos de la promoción para el modal de edición.
    POST: actualiza todos los campos incluyendo días de la semana y productos M2M.
    """
    promo = get_object_or_404(Promocion, pk=id)

    if request.method == "POST":
        datos, errs = _parse_promo_form(request.POST)
        if errs:
            for e in errs:
                messages.error(request, e)
            return redirect("gerente:promociones")

        promo.titulo            = datos["titulo"]
        promo.tipo_descuento    = datos["tipo"]
        promo.valor_descuento   = datos["valor_decimal"]
        promo.cantidad_minima   = datos["cantidad_int"]
        promo.aplicacion        = datos["aplicacion"]
        promo.fecha_inicio      = datos["fecha_inicio"]
        promo.fecha_fin         = datos["fecha_fin"]
        promo.imagen_url        = datos["imagen_url"]
        promo.descripcion_corta = datos["descripcion_corta"]
        promo.orden             = datos["orden"]
        promo.dias_semana       = datos["dias_csv"]
        promo.requiere_todos_productos = datos["requiere_todos"]
        promo.save()

        if datos["ids_aplicables"]:
            promo.productos_aplicables.set(datos["ids_aplicables"])
        else:
            promo.productos_aplicables.clear()
        if datos["ids_beneficiados"]:
            promo.productos_beneficiados.set(datos["ids_beneficiados"])
        else:
            promo.productos_beneficiados.clear()

        messages.success(request, f"Promoción '{datos['titulo']}' actualizada.")
        return redirect("gerente:promociones")

    # GET: devolver JSON con los datos de la promo para cargar en el modal
    data = {
        "id": promo.pk,
        "titulo": promo.titulo,
        "tipo_descuento_id": promo.tipo_descuento_id,
        "aplicacion": promo.aplicacion,
        "valor_descuento": str(promo.valor_descuento) if promo.valor_descuento else "",
        "cantidad_minima": promo.cantidad_minima or "",
        "fecha_inicio": promo.fecha_inicio.strftime("%Y-%m-%dT%H:%M"),
        "fecha_fin": promo.fecha_fin.strftime("%Y-%m-%dT%H:%M"),
        "productos_aplicables": list(promo.productos_aplicables.values_list("id", flat=True)),
        "productos_beneficiados": list(promo.productos_beneficiados.values_list("id", flat=True)),
        "activa": promo.activa,
        "imagen_url": promo.imagen_url or "",
        "descripcion_corta": promo.descripcion_corta or "",
        "orden": promo.orden,
        "dias_semana": (
            [d for d in (promo.dias_semana or "").split(",") if d]
        ),
        "requiere_todos_productos": promo.requiere_todos_productos,
    }
    return JsonResponse(data)


# ─── Empleados ────────────────────────────────────────────────────────────────

@gerente_requerido
def empleados(request):
    """Lista empleados activos e inactivos para la vista de gestión de personal."""
    activos = Empleado.objects.filter(activo=True).order_by("-pk")
    inactivos = Empleado.objects.filter(activo=False).order_by("-pk")
    return render(request, "gerente/empleados.html", {
        "activos": activos, "inactivos": inactivos
    })


@require_POST
@gerente_requerido
def empleados_nuevo(request):
    """Crea un nuevo empleado con usuario y contraseña.

    Solo crea si nombre, usuario y password están presentes;
    silencia el error si faltan campos para no interrumpir flujo.
    """
    nombre = request.POST.get("nombre", "").strip()
    usuario = request.POST.get("usuario", "").strip()
    password = request.POST.get("password", "")
    rol = request.POST.get("rol", "mesero")
    if nombre and usuario and password:
        Empleado.objects.create_user(
            username=usuario, password=password, nombre=nombre, rol=rol
        )
    return redirect("gerente:empleados")


@require_POST
@gerente_requerido
def empleado_toggle(request, id):
    """Activa o desactiva un empleado."""
    emp = get_object_or_404(Empleado, pk=id)
    emp.activo = not emp.activo
    emp.is_active = emp.activo
    emp.save(update_fields=["activo", "is_active"])
    return JsonResponse({"ok": True, "activo": emp.activo})


@require_POST
@gerente_requerido
def empleado_editar(request, id):
    """Actualiza nombre, rol y opcionalmente contraseña de un empleado."""
    emp = get_object_or_404(Empleado, pk=id)
    nombre = request.POST.get("nombre", emp.nombre).strip()
    rol = request.POST.get("rol", emp.rol)
    password = request.POST.get("password", "").strip()
    emp.nombre = nombre
    emp.rol = rol
    if password:
        emp.set_password(password)
    emp.save()
    return redirect("gerente:empleados")


# ─── Reportes ─────────────────────────────────────────────────────────────────

@gerente_requerido
def reportes(request):
    """Vista de reportes básica (legacy). Calcula KPIs, ventas por día y top productos.

    Parámetros GET:
        periodo: "hoy", "mes" o "semana" (default "semana").
    Renderiza gerente/reportes.html con datos JSON para gráficas Chart.js.
    """
    periodo = request.GET.get("periodo", "semana")
    hoy = timezone.now().date()
    if periodo == "hoy":
        desde = hoy
    elif periodo == "mes":
        desde = hoy - timedelta(days=30)
    else:
        desde = hoy - timedelta(days=7)

    base_qs = Pedido.objects.filter(fecha_hora_ingreso__date__gte=desde)

    # BUG #6 FIX: cancelaciones se obtienen de un queryset independiente.
    # Antes se excluían los cancelados al inicio y luego se filtraba sobre el
    # mismo queryset buscando cancelados, obteniendo siempre una lista vacía.
    pedidos = base_qs.exclude(estado="cancelado").prefetch_related(
        "detalles__producto"
    ).select_related("sesion__mesa").order_by("-fecha_hora_ingreso")

    cancelaciones = base_qs.filter(estado="cancelado").select_related("sesion__mesa")

    # KPIs — usar agregación BD en lugar de Python loop (más eficiente y preciso)
    from django.db.models.functions import Coalesce as _Coalesce
    agg    = pedidos.aggregate(t=_Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0.00")))
    total  = agg["t"]                                          # Decimal
    tickets = pedidos.count()
    ticket_promedio = (total / tickets).quantize(Decimal("0.01")) if tickets else Decimal("0.00")

    # Ventas por día — Coalesce garantiza 0 en días sin ventas (evita None en JSON)
    ventas_por_dia_qs = pedidos.annotate(
        fecha=TruncDate("fecha_hora_ingreso")
    ).values("fecha").annotate(
        total=Coalesce(
            Sum("detalles__subtotal_calculado"),
            Decimal("0.00"),
            output_field=DecimalField()
        ),
        tickets=Count("id", distinct=True),
    ).order_by("fecha")

    ventas_por_dia = []
    for row in ventas_por_dia_qs:
        fecha = row.get("fecha") if isinstance(row, dict) else row["fecha"]
        if fecha is None:
            logger.warning("ventas_por_dia: fila con fecha=None ignorada — posible dato corrupto: %r", row)
            continue
        ventas_por_dia.append({
            "dia":     fecha.strftime("%a %d/%m"),
            "total":   float(row["total"]),   # solo float al serializar a JSON
            "tickets": row["tickets"],
        })

    # Productos más vendidos
    top_productos = DetallePedido.objects.filter(
        pedido__fecha_hora_ingreso__date__gte=desde
    ).exclude(pedido__estado="cancelado").values("producto__nombre").annotate(
        total_vendido=Sum("cantidad"),
        ingreso=Sum("subtotal_calculado")
    ).order_by("-total_vendido")[:10]

    return render(request, "gerente/reportes.html", {
        "pedidos": pedidos, "total": total, "periodo": periodo,
        "tickets": tickets, "ticket_promedio": round(ticket_promedio, 2),
        "ventas_por_dia": json.dumps(ventas_por_dia),
        "top_productos": list(top_productos),
        "cancelaciones": cancelaciones,
        "desde": desde,
        "hasta": hoy,
    })


# ─── Exportación de reportes ──────────────────────────────────────────────────

@gerente_requerido
def reporte_exportar(request):
    """GET ?formato=excel|pdf&desde=YYYY-MM-DD&hasta=YYYY-MM-DD"""
    from apps.gerente.reports import exportar_excel, exportar_pdf
    from datetime import date as date_type
    from django.contrib import messages

    fmt = request.GET.get("formato", "excel")
    hoy = timezone.now().date()

    try:
        desde = date_type.fromisoformat(request.GET.get("desde", str(hoy - timedelta(days=7))))
        hasta = date_type.fromisoformat(request.GET.get("hasta", str(hoy)))
    except ValueError:
        messages.error(request, "Fechas inválidas para el reporte.")
        return redirect("gerente:reportes")

    if fmt == "pdf":
        try:
            pdf_bytes = exportar_pdf(desde, hasta)
        except ImportError:
            messages.error(request, "WeasyPrint no está instalado. Ejecuta: pip install weasyprint==62.3 (y reconstruye el contenedor Docker con las dependencias del Dockerfile).")
            return redirect("gerente:reportes")
        except Exception as e:
            messages.error(request, f"Error al generar el PDF: {e}")
            return redirect("gerente:reportes")
        from django.http import HttpResponse
        filename = f"reporte-mochi-{desde}-{hasta}.pdf"
        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    # Default: Excel
    try:
        xlsx_bytes = exportar_excel(desde, hasta)
    except ImportError:
        messages.error(request, "openpyxl no está instalado. Ejecuta: pip install openpyxl==3.1.2")
        return redirect("gerente:reportes")
    except Exception as e:
        messages.error(request, f"Error al generar el Excel: {e}")
        return redirect("gerente:reportes")
    from django.http import HttpResponse
    filename = f"reporte-mochi-{desde}-{hasta}.xlsx"
    resp = HttpResponse(
        xlsx_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ─── Auditoría ────────────────────────────────────────────────────────────────

@gerente_requerido
def auditoria(request):
    """Muestra los últimos 200 registros de auditoría del sistema ordenados por fecha."""
    registros = Auditoria.objects.select_related("empleado", "mesa", "pedido").order_by("-fecha_hora")[:200]
    return render(request, "gerente/reportes.html", {
        "registros": registros, "vista": "auditoria"
    })


# ─── Configuración ────────────────────────────────────────────────────────────

@gerente_requerido
@ensure_csrf_cookie
def configuracion(request):
    """Muestra y guarda la configuración global del sistema.

    POST JSON o form: actualiza claves en el modelo Configuracion mediante
        update_or_create dentro de una transacción atómica.
        Acepta: semaforo_yellow/red, modo_mantenimiento, mensaje_mantenimiento,
        datos del restaurante y credenciales PayPal.
    GET: devuelve el formulario con los valores actuales de cada clave.
    """
    if request.method == "POST":
        from django.db import transaction
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, Exception):
            data = request.POST
        yellow = data.get("yellow")
        red = data.get("red")
        mantenimiento = data.get("modo_mantenimiento")
        mensaje_mant = data.get("mensaje_mantenimiento")
        try:
            with transaction.atomic():
                if yellow is not None:
                    Configuracion.objects.update_or_create(
                        clave="semaforo_yellow", defaults={"valor": str(yellow)}
                    )
                if red is not None:
                    Configuracion.objects.update_or_create(
                        clave="semaforo_red", defaults={"valor": str(red)}
                    )
                if mantenimiento is not None:
                    Configuracion.objects.update_or_create(
                        clave="modo_mantenimiento",
                        defaults={"valor": "true" if str(mantenimiento).lower() in ("true", "1", "yes") else "false"}
                    )
                if mensaje_mant is not None:
                    Configuracion.objects.update_or_create(
                        clave="mensaje_mantenimiento", defaults={"valor": str(mensaje_mant)}
                    )
                # 3.1: flag global de horarios de atención.
                horarios_activos = data.get("horarios_activos")
                if horarios_activos is not None:
                    Configuracion.objects.update_or_create(
                        clave="horarios_activos",
                        defaults={"valor": "true" if str(horarios_activos).lower() in ("true", "1", "yes") else "false"},
                    )
                    # Invalida la caché del middleware para reflejar el cambio sin esperar 30s.
                    try:
                        from config.middleware import HorarioAtencionMiddleware
                        HorarioAtencionMiddleware.invalidate()
                    except Exception:
                        pass
                # Datos del restaurante (para ticket)
                for campo in ("restaurante_nombre", "restaurante_direccion", "restaurante_telefono", "restaurante_rfc"):
                    valor_campo = data.get(campo)
                    if valor_campo is not None:
                        Configuracion.objects.update_or_create(
                            clave=campo, defaults={"valor": str(valor_campo)}
                        )
                # PayPal
                for campo in ("paypal_client_id", "paypal_secret", "paypal_modo"):
                    valor_campo = data.get(campo)
                    if valor_campo is not None:
                        Configuracion.objects.update_or_create(
                            clave=campo, defaults={"valor": str(valor_campo)}
                        )
        except Exception as e:
            return JsonResponse({"ok": False, "error": str(e)[:200]}, status=400)
        return JsonResponse({"ok": True})

    listos_count = Pedido.objects.filter(estado="listo").count()
    yellow_cfg = Configuracion.objects.filter(clave="semaforo_yellow").first()
    red_cfg = Configuracion.objects.filter(clave="semaforo_red").first()
    mant_cfg = Configuracion.objects.filter(clave="modo_mantenimiento").first()
    msg_cfg = Configuracion.objects.filter(clave="mensaje_mantenimiento").first()

    def _cfg(clave, default=""):
        """Helper local: devuelve el valor de una clave de Configuracion o el default."""
        obj = Configuracion.objects.filter(clave=clave).first()
        return obj.valor if obj else default

    # 3.1: estado del flag de horarios + lista para la tabla del panel.
    from apps.gerente.models import HorarioAtencion
    horarios_activos_cfg = Configuracion.objects.filter(clave="horarios_activos").first()
    horarios = HorarioAtencion.objects.all().order_by("dia_semana", "abre")

    return render(request, "gerente/configuracion.html", {
        "listos_count": listos_count,
        "semaforo_yellow": int(yellow_cfg.valor) if yellow_cfg else 8,
        "semaforo_red": int(red_cfg.valor) if red_cfg else 15,
        "modo_mantenimiento": mant_cfg.valor.lower() in ("true", "1") if mant_cfg else False,
        "mensaje_mantenimiento": msg_cfg.valor if msg_cfg else "",
        "restaurante_nombre":    _cfg("restaurante_nombre",    "Mochi Matcha"),
        "restaurante_direccion": _cfg("restaurante_direccion"),
        "restaurante_telefono":  _cfg("restaurante_telefono"),
        "restaurante_rfc":       _cfg("restaurante_rfc"),
        "paypal_client_id": _cfg("paypal_client_id"),
        "paypal_secret":    _cfg("paypal_secret"),
        "paypal_modo":      _cfg("paypal_modo", "sandbox"),
        "horarios_activos": (horarios_activos_cfg.valor.lower() in ("true", "1")) if horarios_activos_cfg else False,
        "horarios":         horarios,
        "dias_semana_choices": HorarioAtencion.DIAS,
    })


# ─── Horarios de atención (3.1) ──────────────────────────────────────────────

@require_POST
@gerente_requerido
def horario_crear(request):
    """
    POST JSON {dia_semana, abre, cierra} → crea un HorarioAtencion.
    Valida: día 0-6, horas con formato HH:MM, abre != cierra.
    """
    from apps.gerente.models import HorarioAtencion
    from datetime import time as dtime
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "JSON inválido"}, status=400)

    try:
        dia = int(data.get("dia_semana", -1))
    except (TypeError, ValueError):
        return JsonResponse({"ok": False, "error": "Día inválido."}, status=400)
    if dia not in range(0, 7):
        return JsonResponse({"ok": False, "error": "Día debe estar entre 0 (lunes) y 6 (domingo)."}, status=400)

    def _parse_hora(txt):
        try:
            h, m = str(txt).split(":")[:2]
            return dtime(int(h), int(m))
        except Exception:
            return None

    abre   = _parse_hora(data.get("abre"))
    cierra = _parse_hora(data.get("cierra"))
    if abre is None or cierra is None:
        return JsonResponse({"ok": False, "error": "Horas inválidas. Formato esperado HH:MM."}, status=400)
    if abre == cierra:
        return JsonResponse({"ok": False, "error": "Abre y cierra no pueden ser iguales."}, status=400)

    h = HorarioAtencion.objects.create(dia_semana=dia, abre=abre, cierra=cierra, activo=True)
    try:
        from config.middleware import HorarioAtencionMiddleware
        HorarioAtencionMiddleware.invalidate()
    except Exception:
        pass
    return JsonResponse({
        "ok": True,
        "id": h.pk,
        "dia_semana": h.dia_semana,
        "dia_display": h.get_dia_semana_display(),
        "abre": h.abre.strftime("%H:%M"),
        "cierra": h.cierra.strftime("%H:%M"),
        "activo": h.activo,
    })


@require_POST
@gerente_requerido
def horario_toggle(request, id):
    """POST → invierte el flag `activo` del horario."""
    from apps.gerente.models import HorarioAtencion
    h = get_object_or_404(HorarioAtencion, pk=id)
    h.activo = not h.activo
    h.save(update_fields=["activo"])
    try:
        from config.middleware import HorarioAtencionMiddleware
        HorarioAtencionMiddleware.invalidate()
    except Exception:
        pass
    return JsonResponse({"ok": True, "activo": h.activo})


@require_POST
@gerente_requerido
def horario_eliminar(request, id):
    """POST → elimina el horario."""
    from apps.gerente.models import HorarioAtencion
    h = get_object_or_404(HorarioAtencion, pk=id)
    h.delete()
    try:
        from config.middleware import HorarioAtencionMiddleware
        HorarioAtencionMiddleware.invalidate()
    except Exception:
        pass
    return JsonResponse({"ok": True})

# ─── Reportes Unificados ─────────────────────────────────────────────────────

@gerente_requerido
def reportes_avanzados(request):
    """Vista única de todos los reportes: ventas, productos, empleados, operativos, cortes."""
    from django.db.models.functions import TruncWeek, TruncMonth
    from django.db.models import Avg, F, ExpressionWrapper, fields as dj_fields

    tipo  = request.GET.get("tipo", "corte_dia")
    hoy   = timezone.now().date()
    desde_str = request.GET.get("desde", str(hoy))
    hasta_str = request.GET.get("hasta", str(hoy))
    try:
        desde = timezone.datetime.strptime(desde_str, "%Y-%m-%d").date()
        hasta = timezone.datetime.strptime(hasta_str, "%Y-%m-%d").date()
    except ValueError:
        desde = hoy; hasta = hoy

    ctx = {"tipo": tipo, "desde": desde, "hasta": hasta, "hoy": hoy}

    base_qs = (
        Pedido.objects
        .filter(fecha_hora_ingreso__date__gte=desde, fecha_hora_ingreso__date__lte=hasta)
        .exclude(estado="cancelado")
    )

    # ── CORTE DEL DÍA ─────────────────────────────────────────────────────────
    if tipo == "corte_dia":
        # Ventas por hora
        from django.db.models.functions import ExtractHour
        por_hora = list(
            base_qs.annotate(hora=ExtractHour("fecha_hora_ingreso"))
            .values("hora")
            .annotate(
                total=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0.00"), output_field=DecimalField()),
                tickets=Count("id", distinct=True),
            ).order_by("hora")
        )
        # Ventas por método de pago
        from apps.pedidos.models import SolicitudPago as SP
        por_metodo = list(
            SP.objects.filter(
                fecha_hora__date__gte=desde, fecha_hora__date__lte=hasta,
                estado_solicitud__descripcion="procesada"
            ).values("metodo_pago__descripcion")
            .annotate(count=Count("id"), propinas=Coalesce(Sum("propina_sugerida"), Decimal("0")))
            .order_by("-count")
        )
        # KPIs globales
        agg = base_qs.aggregate(
            total=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0.00"), output_field=DecimalField()),
            tickets=Count("id", distinct=True),
        )
        propinas = SP.objects.filter(
            fecha_hora__date__gte=desde, fecha_hora__date__lte=hasta,
            estado_solicitud__descripcion="procesada"
        ).aggregate(p=Coalesce(Sum("propina_sugerida"), Decimal("0")))["p"]
        cancelados = Pedido.objects.filter(
            fecha_hora_ingreso__date__gte=desde, fecha_hora_ingreso__date__lte=hasta,
            estado="cancelado"
        ).count()
        # Top productos del día
        top_dia = list(
            DetallePedido.objects
            .filter(pedido__fecha_hora_ingreso__date__gte=desde, pedido__fecha_hora_ingreso__date__lte=hasta)
            .exclude(pedido__estado="cancelado")
            .values("producto__nombre")
            .annotate(cantidad=Sum("cantidad"), ingreso=Sum("subtotal_calculado"))
            .order_by("-cantidad")[:8]
        )
        ctx.update({
            "kpi_total":     agg["total"],
            "kpi_tickets":   agg["tickets"],
            "kpi_propinas":  propinas,
            "kpi_cancelados": cancelados,
            "kpi_promedio":  (agg["total"] / agg["tickets"]).quantize(Decimal("0.01")) if agg["tickets"] else Decimal("0"),
            "por_hora":      por_hora,
            "por_metodo":    por_metodo,
            "top_dia":       top_dia,
            "chart_labels":  json.dumps([f"{r['hora']:02d}:00" for r in por_hora]),
            "chart_data":    json.dumps([float(r["total"]) for r in por_hora]),
            "chart_tickets": json.dumps([r["tickets"] for r in por_hora]),
        })

    # ── CORTE POR MESERO ──────────────────────────────────────────────────────
    elif tipo == "corte_mesero":
        from apps.pedidos.models import SolicitudPago as SP
        meseros = Empleado.objects.filter(rol__in=("mesero", "gerente", "admin"), activo=True)
        tabla = []
        for emp in meseros:
            qs = base_qs.filter(empleado_entrega=emp)
            agg = qs.aggregate(
                ventas=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0"), output_field=DecimalField()),
                pedidos=Count("id", distinct=True),
            )
            # Propinas de las solicitudes de sus sesiones
            propinas = SP.objects.filter(
                sesion__pedidos__empleado_entrega=emp,
                fecha_hora__date__gte=desde, fecha_hora__date__lte=hasta,
                estado_solicitud__descripcion="procesada",
            ).distinct().aggregate(p=Coalesce(Sum("propina_sugerida"), Decimal("0")))["p"]
            cancelados_emp = Pedido.objects.filter(
                empleado_entrega=emp,
                fecha_hora_ingreso__date__gte=desde, fecha_hora_ingreso__date__lte=hasta,
                estado="cancelado"
            ).count()
            promedio = (agg["ventas"] / agg["pedidos"]).quantize(Decimal("0.01")) if agg["pedidos"] else Decimal("0")
            tabla.append({
                "nombre":      emp.nombre,
                "pedidos":     agg["pedidos"],
                "ventas":      agg["ventas"],
                "propinas":    propinas,
                "cancelados":  cancelados_emp,
                "promedio":    promedio,
            })
        tabla.sort(key=lambda x: x["ventas"], reverse=True)
        ctx.update({
            "tabla":         tabla,
            "chart_labels":  json.dumps([r["nombre"] for r in tabla]),
            "chart_data":    json.dumps([float(r["ventas"]) for r in tabla]),
            "chart_propinas":json.dumps([float(r["propinas"]) for r in tabla]),
        })

    # ── VENTAS POR DÍA ────────────────────────────────────────────────────────
    elif tipo == "ventas_dia":
        data = list(
            base_qs.annotate(fecha=TruncDate("fecha_hora_ingreso"))
            .values("fecha")
            .annotate(
                total=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0.00"), output_field=DecimalField()),
                tickets=Count("id", distinct=True),
            ).order_by("fecha")
        )
        ctx.update({
            "chart_labels":  json.dumps([str(r["fecha"]) for r in data]),
            "chart_data":    json.dumps([float(r["total"]) for r in data]),
            "chart_tickets": json.dumps([r["tickets"] for r in data]),
            "tabla": data,
        })

    # ── VENTAS POR SEMANA ─────────────────────────────────────────────────────
    elif tipo == "ventas_semana":
        data = list(
            base_qs.annotate(semana=TruncWeek("fecha_hora_ingreso"))
            .values("semana").annotate(
                total=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0.00"), output_field=DecimalField()),
                tickets=Count("id", distinct=True),
            ).order_by("semana")
        )
        ctx.update({
            "chart_labels":  json.dumps([str(r["semana"])[:10] if r["semana"] else "" for r in data]),
            "chart_data":    json.dumps([float(r["total"]) for r in data]),
            "chart_tickets": json.dumps([r["tickets"] for r in data]),
            "tabla": data,
        })

    # ── VENTAS POR MES ────────────────────────────────────────────────────────
    elif tipo == "ventas_mes":
        data = list(
            base_qs.annotate(mes=TruncMonth("fecha_hora_ingreso"))
            .values("mes").annotate(
                total=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0.00"), output_field=DecimalField()),
                tickets=Count("id", distinct=True),
            ).order_by("mes")
        )
        for r in data:
            r["promedio"] = (r["total"] / r["tickets"]).quantize(Decimal("0.01")) if r["tickets"] else Decimal("0")
        ctx.update({
            "chart_labels":  json.dumps([str(r["mes"])[:7] if r["mes"] else "" for r in data]),
            "chart_data":    json.dumps([float(r["total"]) for r in data]),
            "tabla": data,
        })

    # ── MÉTODO DE PAGO ────────────────────────────────────────────────────────
    elif tipo == "ventas_metodo":
        from apps.pedidos.models import SolicitudPago as SP
        data = list(
            SP.objects.filter(
                fecha_hora__date__gte=desde, fecha_hora__date__lte=hasta,
                estado_solicitud__descripcion="procesada"
            ).values("metodo_pago__descripcion")
            .annotate(count=Count("id"), propinas=Coalesce(Sum("propina_sugerida"), Decimal("0")))
            .order_by("-count")
        )
        ctx.update({
            "chart_labels": json.dumps([r["metodo_pago__descripcion"] or "Sin método" for r in data]),
            "chart_data":   json.dumps([r["count"] for r in data]),
            "tabla": data,
        })

    # ── TOP PRODUCTOS MÁS VENDIDOS ────────────────────────────────────────────
    elif tipo == "top_productos":
        data = list(
            DetallePedido.objects
            .filter(pedido__fecha_hora_ingreso__date__gte=desde, pedido__fecha_hora_ingreso__date__lte=hasta)
            .exclude(pedido__estado="cancelado")
            .values("producto__nombre")
            .annotate(cantidad=Sum("cantidad"), ingreso=Sum("subtotal_calculado"))
            .order_by("-cantidad")[:15]
        )
        ctx.update({
            "chart_labels":  json.dumps([r["producto__nombre"] for r in data]),
            "chart_data":    json.dumps([r["cantidad"] for r in data]),
            "chart_ingreso": json.dumps([float(r["ingreso"] or 0) for r in data]),
            "tabla": data,
        })

    # ── TOP PRODUCTOS MENOS VENDIDOS ──────────────────────────────────────────
    elif tipo == "menos_vendidos":
        from apps.menu.models import Producto as _P
        # Construir mapa de vendidos para lookup O(1); luego iterar el catálogo completo
        vendidos = {
            r["producto__nombre"]: r
            for r in DetallePedido.objects
            .filter(pedido__fecha_hora_ingreso__date__gte=desde, pedido__fecha_hora_ingreso__date__lte=hasta)
            .exclude(pedido__estado="cancelado")
            .values("producto__nombre")
            .annotate(cantidad=Sum("cantidad"), ingreso=Sum("subtotal_calculado"))
        }
        data = []
        # Incluir productos con cantidad=0 (no se vendieron nada en el período)
        for nombre in _P.objects.filter(disponible=True).values_list("nombre", flat=True):
            if nombre in vendidos:
                data.append({"producto__nombre": nombre, "cantidad": vendidos[nombre]["cantidad"], "ingreso": vendidos[nombre]["ingreso"]})
            else:
                data.append({"producto__nombre": nombre, "cantidad": 0, "ingreso": Decimal("0")})
        data.sort(key=lambda x: x["cantidad"])
        data = data[:15]
        ctx.update({
            "chart_labels":  json.dumps([r["producto__nombre"] for r in data]),
            "chart_data":    json.dumps([r["cantidad"] for r in data]),
            "chart_ingreso": json.dumps([float(r["ingreso"] or 0) for r in data]),
            "tabla": data,
        })

    # ── VENTAS POR CATEGORÍA ──────────────────────────────────────────────────
    elif tipo == "ventas_categoria":
        data = list(
            DetallePedido.objects
            .filter(pedido__fecha_hora_ingreso__date__gte=desde, pedido__fecha_hora_ingreso__date__lte=hasta)
            .exclude(pedido__estado="cancelado")
            .values("producto__categoria__nombre")
            .annotate(total=Sum("subtotal_calculado"), cantidad=Sum("cantidad"))
            .order_by("-total")
        )
        ctx.update({
            "chart_labels": json.dumps([r["producto__categoria__nombre"] or "Sin categoría" for r in data]),
            "chart_data":   json.dumps([float(r["total"] or 0) for r in data]),
            "tabla": data,
        })

    # ── DESCUENTOS/PROMOS ─────────────────────────────────────────────────────
    elif tipo == "descuentos_productos":
        data = list(
            DetallePedido.objects
            .filter(pedido__fecha_hora_ingreso__date__gte=desde, pedido__fecha_hora_ingreso__date__lte=hasta)
            .exclude(pedido__estado="cancelado").filter(promocion__isnull=False)
            .values("producto__nombre", "promocion__titulo")
            .annotate(aplicaciones=Count("id"), subtotal=Sum("subtotal_calculado"))
            .order_by("-aplicaciones")[:20]
        )
        ctx["tabla"] = data

    # ── COMPARATIVA EMPLEADOS ─────────────────────────────────────────────────
    elif tipo == "comparativa_empleados":
        meseros = Empleado.objects.filter(rol__in=("mesero", "gerente", "admin"), activo=True)
        data = []
        for emp in meseros:
            agg = base_qs.filter(empleado_entrega=emp).aggregate(
                ventas=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0"), output_field=DecimalField()),
                pedidos=Count("id", distinct=True),
            )
            data.append({"nombre": emp.nombre, "pedidos": agg["pedidos"], "ventas": agg["ventas"]})
        data.sort(key=lambda x: x["ventas"], reverse=True)
        ctx.update({
            "tabla":         data,
            "chart_labels":  json.dumps([r["nombre"] for r in data]),
            "chart_data":    json.dumps([float(r["ventas"]) for r in data]),
            "chart_pedidos": json.dumps([r["pedidos"] for r in data]),
        })

    # ── MESAS UTILIZADAS ──────────────────────────────────────────────────────
    elif tipo == "mesas_utilizadas":
        data = list(
            Pedido.objects
            .filter(fecha_hora_ingreso__date__gte=desde, fecha_hora_ingreso__date__lte=hasta)
            .exclude(estado="cancelado")
            .values("sesion__mesa__numero_mesa")
            .annotate(
                pedidos=Count("id", distinct=True),
                ventas=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0"), output_field=DecimalField())
            ).order_by("-pedidos")
        )
        ctx.update({
            "chart_labels": json.dumps([f"Mesa {r['sesion__mesa__numero_mesa']}" for r in data]),
            "chart_data":   json.dumps([r["pedidos"] for r in data]),
            "tabla": data,
        })

    # ── TIEMPO DE ATENCIÓN ────────────────────────────────────────────────────
    elif tipo == "tiempo_atencion":
        pedidos_con_entrega = (
            Pedido.objects
            .filter(fecha_hora_ingreso__date__gte=desde, fecha_hora_ingreso__date__lte=hasta,
                    fecha_hora_entrega__isnull=False)
            .exclude(estado="cancelado")
            # Calcular duración como campo anotado para poder hacer Avg en BD
            .annotate(dur=ExpressionWrapper(
                F("fecha_hora_entrega") - F("fecha_hora_ingreso"),
                output_field=dj_fields.DurationField()
            ))
        )
        from django.db.models import Avg
        global_avg = pedidos_con_entrega.aggregate(avg=Avg("dur"))["avg"]
        # Convertir timedelta a minutos para presentación
        global_min = round(global_avg.total_seconds() / 60, 1) if global_avg else 0
        meseros_t = []
        for emp in Empleado.objects.filter(rol__in=("mesero","gerente","admin"), activo=True):
            qs = pedidos_con_entrega.filter(empleado_entrega=emp)
            avg = qs.aggregate(avg=Avg("dur"))["avg"]
            meseros_t.append({"nombre": emp.nombre, "pedidos": qs.count(), "tiempo_min": round(avg.total_seconds()/60,1) if avg else 0})
        ctx.update({
            "global_min":    global_min,
            "tabla":         meseros_t,
            "chart_labels":  json.dumps([r["nombre"] for r in meseros_t]),
            "chart_data":    json.dumps([r["tiempo_min"] for r in meseros_t]),
        })

    # ── CANCELACIONES ─────────────────────────────────────────────────────────
    elif tipo == "cancelaciones":
        cancelados = (
            Pedido.objects
            .filter(fecha_hora_ingreso__date__gte=desde, fecha_hora_ingreso__date__lte=hasta, estado="cancelado")
            .select_related("sesion__mesa", "empleado_entrega")
            .order_by("-fecha_hora_ingreso")
        )
        por_mesero = list(
            cancelados.values("empleado_entrega__nombre")
            .annotate(cantidad=Count("id")).order_by("-cantidad")
        )
        ctx.update({"cancelados": cancelados, "por_mesero": por_mesero, "total_cancelaciones": cancelados.count()})

    return render(request, "gerente/reportes_avanzados.html", ctx)



@gerente_requerido
def reportes_avanzados_exportar(request):
    """Exporta el reporte avanzado actual a Excel.

    Parámetros GET: tipo, desde (YYYY-MM-DD), hasta (YYYY-MM-DD).
    Genera las filas según el tipo de reporte (ventas_dia, top_productos,
    cancelaciones, etc.) y devuelve el .xlsx como descarga.
    """
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        from django.contrib import messages
        messages.error(request, "openpyxl no instalado.")
        return redirect("gerente:reportes_avanzados")

    tipo = request.GET.get("tipo", "ventas_dia")
    hoy = timezone.now().date()
    desde_str = request.GET.get("desde", str(hoy - timedelta(days=30)))
    hasta_str = request.GET.get("hasta", str(hoy))
    try:
        desde = timezone.datetime.strptime(desde_str, "%Y-%m-%d").date()
        hasta = timezone.datetime.strptime(hasta_str, "%Y-%m-%d").date()
    except ValueError:
        desde = hoy - timedelta(days=30)
        hasta = hoy

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tipo[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="3A5F3A")

    base_pedidos = (
        Pedido.objects
        .filter(fecha_hora_ingreso__date__gte=desde, fecha_hora_ingreso__date__lte=hasta)
        .exclude(estado="cancelado")
    )

    # Construir datos según tipo (similar a reportes_avanzados)
    if tipo in ("ventas_dia", "ventas_periodo"):
        from django.db.models.functions import TruncDate as _TD
        rows = list(
            base_pedidos.annotate(fecha=_TD("fecha_hora_ingreso"))
            .values("fecha")
            .annotate(
                total=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0.00"), output_field=DecimalField()),
                tickets=Count("id", distinct=True),
            ).order_by("fecha")
        )
        headers = ["Fecha", "Ventas ($)", "Tickets"]
        ws.append(headers)
        for i, h in enumerate(headers, 1):
            cell = ws.cell(1, i)
            cell.font = header_font
            cell.fill = header_fill
        for r in rows:
            ws.append([str(r["fecha"]), float(r["total"]), r["tickets"]])

    elif tipo == "top_productos" or tipo == "menos_vendidos":
        rows = list(
            DetallePedido.objects
            .filter(pedido__fecha_hora_ingreso__date__gte=desde, pedido__fecha_hora_ingreso__date__lte=hasta)
            .exclude(pedido__estado="cancelado")
            .values("producto__nombre")
            .annotate(cantidad=Sum("cantidad"), ingreso=Sum("subtotal_calculado"))
            .order_by("-cantidad" if tipo == "top_productos" else "cantidad")[:10]
        )
        headers = ["Producto", "Unidades vendidas", "Ingreso ($)"]
        ws.append(headers)
        for i, h in enumerate(headers, 1):
            cell = ws.cell(1, i)
            cell.font = header_font
            cell.fill = header_fill
        for r in rows:
            ws.append([r["producto__nombre"], r["cantidad"], float(r["ingreso"] or 0)])

    elif tipo == "cancelaciones":
        cancelados = list(
            Pedido.objects
            .filter(fecha_hora_ingreso__date__gte=desde, fecha_hora_ingreso__date__lte=hasta, estado="cancelado")
            .select_related("sesion__mesa", "empleado_entrega")
            .order_by("-fecha_hora_ingreso")
        )
        headers = ["Pedido #", "Fecha", "Mesa", "Mesero", "Motivo"]
        ws.append(headers)
        for i, h in enumerate(headers, 1):
            cell = ws.cell(1, i)
            cell.font = header_font
            cell.fill = header_fill
        for p in cancelados:
            ws.append([
                p.pk,
                str(p.fecha_hora_ingreso.date()),
                p.sesion.mesa.numero_mesa if p.sesion and p.sesion.mesa else "",
                p.empleado_entrega.nombre if p.empleado_entrega else "",
                p.motivo_cancelacion,
            ])
    else:
        ws.append(["Reporte", tipo, "Desde", str(desde), "Hasta", str(hasta)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from django.http import HttpResponse
    filename = f"reporte_{tipo}_{desde}_{hasta}.xlsx"
    resp = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ─── Exportación personalizada (secciones + gráficas) ────────────────────────

@gerente_requerido
def reportes_exportar_custom(request):
    """Exporta un reporte con secciones seleccionadas por el usuario."""
    from apps.gerente.reports import exportar_excel_custom, exportar_pdf_custom, SECCIONES_VALIDAS

    formato   = request.GET.get("formato", "excel")
    secciones = [s for s in request.GET.getlist("secciones") if s in SECCIONES_VALIDAS]

    hoy = timezone.now().date()
    try:
        desde = timezone.datetime.strptime(request.GET.get("desde", str(hoy)), "%Y-%m-%d").date()
        hasta = timezone.datetime.strptime(request.GET.get("hasta", str(hoy)), "%Y-%m-%d").date()
    except ValueError:
        desde = hoy
        hasta = hoy

    if not secciones:
        messages.error(request, "Selecciona al menos una sección para exportar.")
        return redirect("gerente:reportes_avanzados")

    from django.http import HttpResponse

    if formato == "pdf":
        try:
            content = exportar_pdf_custom(desde, hasta, secciones)
            ct      = "application/pdf"
            ext     = "pdf"
        except (ImportError, RuntimeError) as e:
            messages.error(request, str(e))
            return redirect("gerente:reportes_avanzados")
    else:
        try:
            content = exportar_excel_custom(desde, hasta, secciones)
            ct      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ext     = "xlsx"
        except (ImportError, RuntimeError) as e:
            messages.error(request, str(e))
            return redirect("gerente:reportes_avanzados")

    resp = HttpResponse(content, content_type=ct)
    resp["Content-Disposition"] = f'attachment; filename="reporte-mochi-{desde}-{hasta}.{ext}"'
    return resp


# ─── Upload de imágenes (drag-and-drop) ───────────────────────────────────────
@gerente_requerido
@require_POST
def upload_imagen(request):
    """
    Recibe un archivo de imagen vía multipart/form-data y lo guarda en
    MEDIA_ROOT/images/ con un nombre único. Devuelve la ruta relativa (p.ej.
    /media/images/abc123.jpg) que el frontend pone en el campo imagen_url.
    Solo guarda la RUTA en la BD, el archivo vive en el filesystem.
    """
    import os, uuid
    from django.conf import settings

    archivo = request.FILES.get("imagen") or request.FILES.get("file")
    if not archivo:
        return JsonResponse({"ok": False, "error": "No se recibió archivo."}, status=400)

    # Validar tipo
    EXT_VALIDAS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
    nombre_original = archivo.name or "imagen"
    ext = os.path.splitext(nombre_original)[1].lower()
    if ext not in EXT_VALIDAS:
        return JsonResponse({
            "ok": False,
            "error": f"Formato no permitido. Usa: {', '.join(sorted(EXT_VALIDAS))}"
        }, status=400)

    # Validar tamaño (máx 5 MB)
    if archivo.size > 5 * 1024 * 1024:
        return JsonResponse({"ok": False, "error": "Imagen mayor a 5 MB."}, status=400)

    # Carpeta destino
    destino_dir = os.path.join(settings.MEDIA_ROOT, "images")
    os.makedirs(destino_dir, exist_ok=True)

    # Nombre único para evitar colisiones / problemas con espacios
    slug = uuid.uuid4().hex[:12]
    nombre_final = f"{slug}{ext}"
    ruta_abs = os.path.join(destino_dir, nombre_final)
    with open(ruta_abs, "wb") as f:
        for chunk in archivo.chunks():
            f.write(chunk)

    ruta_relativa = f"{settings.MEDIA_URL}images/{nombre_final}"
    return JsonResponse({"ok": True, "url": ruta_relativa})


@gerente_requerido
@require_GET
def listar_imagenes(request):
    """
    Lista imágenes en MEDIA_ROOT/images con metadatos para la galería.
    Soporta filtros por nombre (`q`) y por fecha (`desde`, `hasta` en YYYY-MM-DD).
    Devuelve [{nombre, url, tamano, mtime, fecha}, ...] ordenado por mtime desc.
    """
    import os
    from datetime import datetime, date
    from django.conf import settings

    q     = (request.GET.get("q") or "").strip().lower()
    desde = (request.GET.get("desde") or "").strip()
    hasta = (request.GET.get("hasta") or "").strip()

    def _parse(s):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            return None

    d1 = _parse(desde)
    d2 = _parse(hasta)

    EXT_VALIDAS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
    dir_imgs = os.path.join(settings.MEDIA_ROOT, "images")
    items = []
    if os.path.isdir(dir_imgs):
        for fname in os.listdir(dir_imgs):
            ext = os.path.splitext(fname)[1].lower()
            if ext not in EXT_VALIDAS:
                continue
            if q and q not in fname.lower():
                continue
            ruta_abs = os.path.join(dir_imgs, fname)
            try:
                stat = os.stat(ruta_abs)
            except OSError:
                continue
            fecha_archivo = date.fromtimestamp(stat.st_mtime)
            if d1 and fecha_archivo < d1:
                continue
            if d2 and fecha_archivo > d2:
                continue
            items.append({
                "nombre": fname,
                "url": f"{settings.MEDIA_URL}images/{fname}",
                "tamano": stat.st_size,
                "mtime": stat.st_mtime,
                "fecha": fecha_archivo.isoformat(),
            })
    items.sort(key=lambda x: x["mtime"], reverse=True)
    return JsonResponse({"ok": True, "items": items})


@gerente_requerido
@require_POST
def eliminar_imagen(request):
    """
    Elimina un archivo de MEDIA_ROOT/images. Recibe `nombre` (sólo el filename,
    sin path). Bloquea path traversal y solo opera dentro de MEDIA_ROOT/images.
    """
    import os
    from django.conf import settings

    try:
        data = json.loads(request.body or b"{}")
    except (json.JSONDecodeError, Exception):
        data = request.POST
    nombre = (data.get("nombre") or "").strip()
    if not nombre or "/" in nombre or "\\" in nombre or ".." in nombre:
        return JsonResponse({"ok": False, "error": "Nombre inválido."}, status=400)

    dir_imgs = os.path.realpath(os.path.join(settings.MEDIA_ROOT, "images"))
    ruta_abs = os.path.realpath(os.path.join(dir_imgs, nombre))
    if not ruta_abs.startswith(dir_imgs + os.sep):
        return JsonResponse({"ok": False, "error": "Ruta fuera de media/images."}, status=400)
    if not os.path.isfile(ruta_abs):
        return JsonResponse({"ok": False, "error": "Archivo no encontrado."}, status=404)

    # Advertencia si la imagen está en uso (no impedir, solo informar)
    url_relativa = f"{settings.MEDIA_URL}images/{nombre}"
    en_uso_prod  = Producto.objects.filter(imagen_url=url_relativa).count()
    en_uso_promo = Promocion.objects.filter(imagen_url=url_relativa).count()

    try:
        os.remove(ruta_abs)
    except OSError as e:
        return JsonResponse({"ok": False, "error": f"No se pudo eliminar: {e}"}, status=500)

    # Limpiar referencias en BD (dejar imagen_url vacío para que use fallback)
    if en_uso_prod:
        Producto.objects.filter(imagen_url=url_relativa).update(imagen_url="")
    if en_uso_promo:
        Promocion.objects.filter(imagen_url=url_relativa).update(imagen_url="")

    return JsonResponse({
        "ok": True,
        "limpiados": {"productos": en_uso_prod, "promociones": en_uso_promo},
    })


