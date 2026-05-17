"""
gerente/reports.py
Lógica de generación de reportes en Excel y PDF para Mochi Matcha.

Dependencias requeridas (añadir a requirements.txt):
    openpyxl==3.1.2
    weasyprint==62.3
"""
from __future__ import annotations

import io
from datetime import date, timedelta
from decimal import Decimal
from typing import Iterable

from django.db.models import Sum, Count, DecimalField
from django.db.models.functions import TruncDate, Coalesce
from django.template.loader import render_to_string
from django.utils import timezone

from apps.pedidos.models import Pedido, DetallePedido


def _make_naive(dt):
    """Convierte un datetime aware a naive (local) para openpyxl, que no soporta tzinfo."""
    if dt is None:
        return None
    if hasattr(dt, "tzinfo") and dt.tzinfo is not None:
        return timezone.localtime(dt).replace(tzinfo=None)
    return dt


# ──────────────────────────────────────────────────────────────────────────────
# Helpers de datos
# ──────────────────────────────────────────────────────────────────────────────

def _pedidos_periodo(desde: date, hasta: date):
    """Devuelve queryset de pedidos no cancelados en el rango [desde, hasta]."""
    return (
        Pedido.objects
        .filter(fecha_hora_ingreso__date__gte=desde, fecha_hora_ingreso__date__lte=hasta)
        .exclude(estado="cancelado")
        .prefetch_related("detalles__producto")
        .select_related("sesion__mesa")
    )


def _ventas_por_dia(pedidos):
    """Agrupa el queryset de pedidos por día y devuelve lista con total y tickets por fecha.

    Parámetros:
        pedidos: queryset de Pedido ya filtrado y sin cancelados.
    Retorna:
        list[dict] con claves: fecha (date), total (Decimal), tickets (int).
    """
    return list(
        pedidos
        .annotate(fecha=TruncDate("fecha_hora_ingreso"))
        .values("fecha")
        .annotate(
            # Coalesce evita None en días con subtotales NULL (p.ej. pedidos sin detalles)
            total=Coalesce(
            Sum("detalles__subtotal_calculado"),
            Decimal("0.00"),
            output_field=DecimalField()
        ),
            tickets=Count("id", distinct=True),
        )
        .order_by("fecha")
    )


def _top_productos(desde: date, hasta: date, limit: int = 20):
    """Devuelve los productos más vendidos en el período, ordenados por cantidad desc.

    Parámetros:
        desde: fecha de inicio (inclusive).
        hasta: fecha de fin (inclusive).
        limit: número máximo de productos a devolver (default 20).
    Retorna:
        list[dict] con claves: producto__nombre, cantidad, ingreso (Decimal).
    """
    return list(
        DetallePedido.objects
        .filter(pedido__fecha_hora_ingreso__date__gte=desde,
                pedido__fecha_hora_ingreso__date__lte=hasta)
        .exclude(pedido__estado="cancelado")
        .values("producto__nombre")
        .annotate(cantidad=Sum("cantidad"), ingreso=Sum("subtotal_calculado"))
        .order_by("-cantidad")[:limit]
    )


def _cancelaciones(desde: date, hasta: date):
    """Devuelve lista de pedidos cancelados en el período con datos mínimos para el reporte.

    Parámetros:
        desde: fecha de inicio (inclusive).
        hasta: fecha de fin (inclusive).
    Retorna:
        list[dict] con claves: id, fecha_hora_ingreso, motivo_cancelacion,
        sesion__mesa__numero_mesa, sesion__alias.
    """
    return list(
        Pedido.objects
        .filter(estado="cancelado",
                fecha_hora_ingreso__date__gte=desde,
                fecha_hora_ingreso__date__lte=hasta)
        .select_related("sesion__mesa")
        .values("id", "fecha_hora_ingreso", "motivo_cancelacion",
                "sesion__mesa__numero_mesa", "sesion__alias")
    )


def _rendimiento_meseros(desde: date, hasta: date) -> list:
    """Devuelve rendimiento por mesero para el período dado."""
    from decimal import Decimal
    from django.db.models import DecimalField
    from django.db.models.functions import Coalesce
    from apps.accounts.models import Empleado
    from apps.pedidos.models import Pedido

    resultado = []
    meseros = Empleado.objects.filter(rol__in=("mesero", "gerente", "admin"), activo=True)
    for emp in meseros:
        # Una query por mesero: aceptable porque el número de empleados es pequeño
        qs = (
            Pedido.objects
            .filter(fecha_hora_ingreso__date__gte=desde, fecha_hora_ingreso__date__lte=hasta,
                    empleado_entrega=emp)
            .exclude(estado="cancelado")
        )
        agg = qs.aggregate(
            ventas=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0.00"), output_field=DecimalField()),
            pedidos=Count("id", distinct=True),
        )
        resultado.append({
            "nombre": emp.nombre,
            "pedidos": agg["pedidos"],
            "ventas": agg["ventas"],
        })
    # Ordenar de mayor a menor ventas para que el ranking sea inmediato en el reporte
    return sorted(resultado, key=lambda x: x["ventas"], reverse=True)


def get_report_data(desde: date, hasta: date) -> dict:
    """Devuelve un dict con todos los datos necesarios para el reporte.

    Los campos 'total' y 'ticket_promedio' se devuelven como Decimal para
    mantener precisión. Se convierten a float/str en el punto de serialización
    (JSON en la vista, o celdas de Excel en exportar_excel).
    """
    pedidos  = _pedidos_periodo(desde, hasta)
    agg      = pedidos.aggregate(
        t=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0.00"))
    )
    total    = agg["t"]                                    # Decimal
    tickets  = pedidos.count()
    ticket_promedio = (total / tickets).quantize(Decimal("0.01")) if tickets else Decimal("0.00")
    return {
        "desde":           desde,
        "hasta":           hasta,
        "total":           total,           # Decimal — float solo al serializar
        "tickets":         tickets,
        "ticket_promedio": ticket_promedio, # Decimal
        "ventas_por_dia":  _ventas_por_dia(pedidos),
        "top_productos":   _top_productos(desde, hasta),
        "cancelaciones":   _cancelaciones(desde, hasta),
        "rendimiento_meseros": _rendimiento_meseros(desde, hasta),
    }


# ──────────────────────────────────────────────────────────────────────────────
# Exportación Excel
# ──────────────────────────────────────────────────────────────────────────────
# Exportación Excel
# ──────────────────────────────────────────────────────────────────────────────

def exportar_excel(desde: date, hasta: date) -> bytes:
    """Genera un archivo .xlsx y devuelve sus bytes.

    Raises:
        ImportError: si openpyxl no está instalado.
        RuntimeError: si la generación falla por cualquier otro motivo.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl no está instalado. Ejecuta: pip install openpyxl==3.1.2"
        )

    try:
        data = get_report_data(desde, hasta)
    except Exception as exc:
        raise RuntimeError(f"Error al obtener datos del reporte: {exc}") from exc

    try:
        wb = openpyxl.Workbook()

        # ── Hoja 1: Resumen ──────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Resumen"

        VERDE = "2D6A4F"
        VERDE_CLARO = "D8F3DC"
        GRIS = "F8F9FA"

        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor=VERDE)
        subheader_font = Font(name="Calibri", bold=True, size=10)
        subheader_fill = PatternFill("solid", fgColor=VERDE_CLARO)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _header(ws, row, col, text, span=1):
            cell = ws.cell(row=row, column=col, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            if span > 1:
                ws.merge_cells(start_row=row, start_column=col,
                               end_row=row, end_column=col + span - 1)
            return cell

        def _cell(ws, row, col, value, fmt=None, bold=False):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if fmt:
                cell.number_format = fmt
            if bold:
                cell.font = Font(bold=True)
            return cell

        # Título
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = f"Reporte Mochi Matcha  |  {desde.strftime('%d/%m/%Y')} – {hasta.strftime('%d/%m/%Y')}"
        title_cell.font = Font(name="Calibri", bold=True, size=14, color=VERDE)
        title_cell.alignment = center
        ws.row_dimensions[1].height = 28

        # KPIs
        _header(ws, 3, 1, "INDICADOR", 2)
        _header(ws, 3, 3, "VALOR", 2)
        kpis = [
            ("Ventas totales", f"${data['total']:,.2f}"),
            ("Pedidos procesados", data["tickets"]),
            ("Ticket promedio", f"${data['ticket_promedio']:,.2f}"),
            ("Cancelaciones", len(data["cancelaciones"])),
        ]
        for i, (k, v) in enumerate(kpis, start=4):
            fill = PatternFill("solid", fgColor=GRIS if i % 2 == 0 else "FFFFFF")
            for col in range(1, 5):
                c = ws.cell(row=i, column=col)
                c.fill = fill
                c.border = border
            ws.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
            ws.cell(row=i, column=3, value=v)
            ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=4)

        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 22

        # ── Hoja 2: Ventas por día ───────────────────────────────────────────────
        ws2 = wb.create_sheet("Ventas por día")
        _header(ws2, 1, 1, "FECHA")
        _header(ws2, 1, 2, "PEDIDOS")
        _header(ws2, 1, 3, "TOTAL ($)")
        for i, row in enumerate(data["ventas_por_dia"], start=2):
            _cell(ws2, i, 1, _make_naive(row.get("fecha")), "DD/MM/YYYY")
            _cell(ws2, i, 2, row.get("tickets", 0))
            _cell(ws2, i, 3, float(row.get("total") or 0), "#,##0.00")
        for col in [1, 2, 3]:
            ws2.column_dimensions[get_column_letter(col)].width = 20

        # ── Hoja 3: Top productos ────────────────────────────────────────────────
        ws3 = wb.create_sheet("Productos")
        _header(ws3, 1, 1, "#")
        _header(ws3, 1, 2, "PRODUCTO")
        _header(ws3, 1, 3, "CANTIDAD VENDIDA")
        _header(ws3, 1, 4, "INGRESOS ($)")
        for i, row in enumerate(data["top_productos"], start=2):
            _cell(ws3, i, 1, i - 1)
            _cell(ws3, i, 2, row.get("producto__nombre", "—"))
            _cell(ws3, i, 3, int(row.get("cantidad") or 0))
            _cell(ws3, i, 4, float(row.get("ingreso") or 0), "#,##0.00")
        ws3.column_dimensions["A"].width = 6
        ws3.column_dimensions["B"].width = 35
        ws3.column_dimensions["C"].width = 20
        ws3.column_dimensions["D"].width = 20

        # ── Hoja 4: Cancelaciones ────────────────────────────────────────────────
        ws4 = wb.create_sheet("Cancelaciones")
        _header(ws4, 1, 1, "ID")
        _header(ws4, 1, 2, "FECHA")
        _header(ws4, 1, 3, "MESA")
        _header(ws4, 1, 4, "CLIENTE")
        _header(ws4, 1, 5, "MOTIVO")
        for i, row in enumerate(data["cancelaciones"], start=2):
            _cell(ws4, i, 1, row.get("id"))
            _cell(ws4, i, 2, _make_naive(row.get("fecha_hora_ingreso")), "DD/MM/YYYY HH:MM")
            _cell(ws4, i, 3, row.get("sesion__mesa__numero_mesa"))
            _cell(ws4, i, 4, row.get("sesion__alias") or "—")
            _cell(ws4, i, 5, row.get("motivo_cancelacion") or "—")
        for col, w in zip(range(1, 6), [8, 20, 10, 20, 50]):
            ws4.column_dimensions[get_column_letter(col)].width = w

        # ── Hoja 5: Rendimiento Meseros ──────────────────────────────────────────
        ws5 = wb.create_sheet("Rendimiento Meseros")
        _header(ws5, 1, 1, "MESERO")
        _header(ws5, 1, 2, "PEDIDOS ENTREGADOS")
        _header(ws5, 1, 3, "VENTAS ($)")
        for i, row in enumerate(data.get("rendimiento_meseros", []), start=2):
            _cell(ws5, i, 1, row.get("nombre", "—"))
            _cell(ws5, i, 2, row.get("pedidos", 0))
            _cell(ws5, i, 3, float(row.get("ventas") or 0), "#,##0.00")
        ws5.column_dimensions["A"].width = 30
        ws5.column_dimensions["B"].width = 22
        ws5.column_dimensions["C"].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    except Exception as exc:
        raise RuntimeError(f"Error al generar el archivo Excel: {exc}") from exc


# ──────────────────────────────────────────────────────────────────────────────
# Exportación PDF (via WeasyPrint → HTML → PDF)
# ──────────────────────────────────────────────────────────────────────────────

def exportar_pdf(desde: date, hasta: date) -> bytes:
    """Genera un PDF a partir de un template HTML y devuelve bytes.

    Raises:
        ImportError: si weasyprint no está instalado.
        RuntimeError: si la renderización o generación PDF falla.
    """
    try:
        import weasyprint
    except ImportError:
        raise ImportError(
            "weasyprint no está instalado. Ejecuta: pip install weasyprint==62.3 "
            "(y reconstruye la imagen Docker con las dependencias del sistema)."
        )

    try:
        data = get_report_data(desde, hasta)
    except Exception as exc:
        raise RuntimeError(f"Error al obtener datos del reporte: {exc}") from exc

    try:
        html_str = render_to_string("gerente/reporte_pdf.html", {
            "desde": desde,
            "hasta": hasta,
            **data,
        })
        pdf_bytes = weasyprint.HTML(string=html_str).write_pdf()
        return pdf_bytes
    except Exception as exc:
        raise RuntimeError(f"Error al renderizar el PDF: {exc}") from exc


# ──────────────────────────────────────────────────────────────────────────────
# Exportación personalizada (secciones seleccionables + gráficas)
# ──────────────────────────────────────────────────────────────────────────────

SECCIONES_VALIDAS = {
    "corte_dia", "corte_mesero",
    "ventas_dia", "ventas_semana", "ventas_mes", "ventas_metodo",
    "top_productos", "menos_vendidos", "ventas_categoria", "descuentos_productos",
    "comparativa_empleados", "tiempo_atencion",
    "mesas_utilizadas", "cancelaciones",
}

SECCION_TITULOS = {
    "corte_dia":             "Corte del Día",
    "corte_mesero":          "Corte por Mesero",
    "ventas_dia":            "Ventas por Día",
    "ventas_semana":         "Ventas por Semana",
    "ventas_mes":            "Ventas por Mes",
    "ventas_metodo":         "Por Método de Pago",
    "top_productos":         "Productos Más Vendidos",
    "menos_vendidos":        "Productos Menos Vendidos",
    "ventas_categoria":      "Ventas por Categoría",
    "descuentos_productos":  "Promociones Aplicadas",
    "comparativa_empleados": "Comparativa de Empleados",
    "tiempo_atencion":       "Tiempo de Atención",
    "mesas_utilizadas":      "Mesas Utilizadas",
    "cancelaciones":         "Cancelaciones",
}


def _datos_seccion(tipo: str, desde: date, hasta: date) -> dict:
    """Devuelve datos estructurados para un tipo de reporte.

    Calcula KPIs, filas de tabla y datos de gráfica para la sección indicada.
    Es usada tanto por exportar_excel_custom() como por exportar_pdf_custom().

    Parámetros:
        tipo:  clave de sección (ver SECCIONES_VALIDAS).
        desde: fecha de inicio (inclusive).
        hasta: fecha de fin (inclusive).
    Retorna:
        dict con claves: titulo, kpis, tabla_headers, tabla_rows,
        tablas_extra, chart_type, chart_labels, chart_values, chart_title.
    """
    from apps.pedidos.models import Pedido, DetallePedido, SolicitudPago as SP
    from apps.accounts.models import Empleado

    # Queryset base reutilizado por la mayoría de secciones: pedidos no cancelados en el rango
    base_qs = (
        Pedido.objects
        .filter(fecha_hora_ingreso__date__gte=desde, fecha_hora_ingreso__date__lte=hasta)
        .exclude(estado="cancelado")
    )

    titulo        = SECCION_TITULOS.get(tipo, tipo)
    kpis          = []
    tabla_headers = []
    tabla_rows    = []
    tablas_extra  = []
    chart_type    = None
    chart_labels  = []
    chart_values  = []
    chart_title   = ""

    if tipo == "corte_dia":
        from django.db.models.functions import ExtractHour
        por_hora = list(
            base_qs.annotate(hora=ExtractHour("fecha_hora_ingreso"))
            .values("hora")
            .annotate(
                total=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0.00"), output_field=DecimalField()),
                tickets=Count("id", distinct=True),
            ).order_by("hora")
        )
        por_metodo = list(
            SP.objects.filter(
                fecha_hora__date__gte=desde, fecha_hora__date__lte=hasta,
                estado_solicitud__descripcion="procesada"
            ).values("metodo_pago__descripcion")
            .annotate(count=Count("id"), propinas=Coalesce(Sum("propina_sugerida"), Decimal("0")))
            .order_by("-count")
        )
        agg = base_qs.aggregate(
            total=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0.00"), output_field=DecimalField()),
            tickets=Count("id", distinct=True),
        )
        propinas = SP.objects.filter(
            fecha_hora__date__gte=desde, fecha_hora__date__lte=hasta,
            estado_solicitud__descripcion="procesada"
        ).aggregate(p=Coalesce(Sum("propina_sugerida"), Decimal("0")))["p"]
        cancelados = Pedido.objects.filter(
            fecha_hora_ingreso__date__gte=desde,
            fecha_hora_ingreso__date__lte=hasta,
            estado="cancelado"
        ).count()
        top_dia = list(
            DetallePedido.objects
            .filter(pedido__fecha_hora_ingreso__date__gte=desde,
                    pedido__fecha_hora_ingreso__date__lte=hasta)
            .exclude(pedido__estado="cancelado")
            .values("producto__nombre")
            .annotate(cantidad=Sum("cantidad"), ingreso=Sum("subtotal_calculado"))
            .order_by("-cantidad")[:8]
        )
        promedio = (agg["total"] / agg["tickets"]).quantize(Decimal("0.01")) if agg["tickets"] else Decimal("0")
        kpis = [
            ("Ventas totales",  f"${float(agg['total']):,.2f}"),
            ("Pedidos",         str(agg["tickets"])),
            ("Ticket promedio", f"${float(promedio):,.2f}"),
            ("Propinas",        f"${float(propinas):,.2f}"),
            ("Cancelaciones",   str(cancelados)),
        ]
        tabla_headers = ["Hora", "Ventas ($)", "Tickets"]
        tabla_rows    = [[f"{r['hora']:02d}:00", float(r["total"]), r["tickets"]] for r in por_hora]
        tablas_extra  = [
            {
                "titulo":  "Por método de pago",
                "headers": ["Método", "Transacciones", "Propinas ($)"],
                "rows":    [[r["metodo_pago__descripcion"] or "N/A", r["count"], float(r["propinas"])]
                            for r in por_metodo],
            },
            {
                "titulo":  "Top productos del día",
                "headers": ["Producto", "Unidades", "Ingreso ($)"],
                "rows":    [[r["producto__nombre"], r["cantidad"], float(r["ingreso"] or 0)]
                            for r in top_dia],
            },
        ]
        chart_type   = "bar"
        chart_labels = [f"{r['hora']:02d}:00" for r in por_hora]
        chart_values = [float(r["total"]) for r in por_hora]
        chart_title  = "Ventas por hora"

    elif tipo == "corte_mesero":
        meseros = Empleado.objects.filter(rol__in=("mesero", "gerente", "admin"), activo=True)
        tabla_data = []
        for emp in meseros:
            qs  = base_qs.filter(empleado_entrega=emp)
            agg = qs.aggregate(
                ventas=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0"), output_field=DecimalField()),
                pedidos=Count("id", distinct=True),
            )
            propinas = SP.objects.filter(
                sesion__pedidos__empleado_entrega=emp,
                fecha_hora__date__gte=desde, fecha_hora__date__lte=hasta,
                estado_solicitud__descripcion="procesada",
            ).distinct().aggregate(p=Coalesce(Sum("propina_sugerida"), Decimal("0")))["p"]
            cancelados_emp = Pedido.objects.filter(
                empleado_entrega=emp,
                fecha_hora_ingreso__date__gte=desde,
                fecha_hora_ingreso__date__lte=hasta,
                estado="cancelado"
            ).count()
            promedio = (agg["ventas"] / agg["pedidos"]).quantize(Decimal("0.01")) if agg["pedidos"] else Decimal("0")
            tabla_data.append({
                "nombre":    emp.nombre,
                "pedidos":   agg["pedidos"],
                "ventas":    float(agg["ventas"]),
                "propinas":  float(propinas),
                "cancelados": cancelados_emp,
                "promedio":  float(promedio),
            })
        tabla_data.sort(key=lambda x: x["ventas"], reverse=True)
        tabla_headers = ["Mesero", "Pedidos", "Ventas ($)", "Propinas ($)", "Cancelados", "Promedio ($)"]
        tabla_rows    = [[r["nombre"], r["pedidos"], r["ventas"], r["propinas"],
                          r["cancelados"], r["promedio"]] for r in tabla_data]
        chart_type   = "barh"
        chart_labels = [r["nombre"] for r in tabla_data]
        chart_values = [r["ventas"] for r in tabla_data]
        chart_title  = "Ventas por mesero"

    elif tipo == "ventas_dia":
        from django.db.models.functions import TruncDate as _TD
        data = list(
            base_qs.annotate(fecha=_TD("fecha_hora_ingreso"))
            .values("fecha")
            .annotate(
                total=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0.00"), output_field=DecimalField()),
                tickets=Count("id", distinct=True),
            ).order_by("fecha")
        )
        tabla_headers = ["Fecha", "Ventas ($)", "Tickets"]
        tabla_rows    = [[str(r["fecha"]), float(r["total"]), r["tickets"]] for r in data]
        chart_type   = "line"
        chart_labels = [str(r["fecha"]) for r in data]
        chart_values = [float(r["total"]) for r in data]
        chart_title  = "Ventas por día"

    elif tipo == "ventas_semana":
        from django.db.models.functions import TruncWeek
        data = list(
            base_qs.annotate(semana=TruncWeek("fecha_hora_ingreso"))
            .values("semana")
            .annotate(
                total=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0.00"), output_field=DecimalField()),
                tickets=Count("id", distinct=True),
            ).order_by("semana")
        )
        tabla_headers = ["Semana", "Ventas ($)", "Tickets"]
        tabla_rows    = [[str(r["semana"])[:10] if r["semana"] else "", float(r["total"]), r["tickets"]]
                         for r in data]
        chart_type   = "line"
        chart_labels = [str(r["semana"])[:10] if r["semana"] else "" for r in data]
        chart_values = [float(r["total"]) for r in data]
        chart_title  = "Ventas por semana"

    elif tipo == "ventas_mes":
        from django.db.models.functions import TruncMonth
        data = list(
            base_qs.annotate(mes=TruncMonth("fecha_hora_ingreso"))
            .values("mes")
            .annotate(
                total=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0.00"), output_field=DecimalField()),
                tickets=Count("id", distinct=True),
            ).order_by("mes")
        )
        for r in data:
            r["promedio"] = float((r["total"] / r["tickets"]).quantize(Decimal("0.01"))) if r["tickets"] else 0.0
        tabla_headers = ["Mes", "Ventas ($)", "Tickets", "Promedio ($)"]
        tabla_rows    = [[str(r["mes"])[:7] if r["mes"] else "", float(r["total"]),
                          r["tickets"], r["promedio"]] for r in data]
        chart_type   = "bar"
        chart_labels = [str(r["mes"])[:7] if r["mes"] else "" for r in data]
        chart_values = [float(r["total"]) for r in data]
        chart_title  = "Ventas por mes"

    elif tipo == "ventas_metodo":
        data = list(
            SP.objects.filter(
                fecha_hora__date__gte=desde, fecha_hora__date__lte=hasta,
                estado_solicitud__descripcion="procesada"
            ).values("metodo_pago__descripcion")
            .annotate(count=Count("id"), propinas=Coalesce(Sum("propina_sugerida"), Decimal("0")))
            .order_by("-count")
        )
        tabla_headers = ["Método de pago", "Transacciones", "Propinas ($)"]
        tabla_rows    = [[r["metodo_pago__descripcion"] or "Sin método", r["count"], float(r["propinas"])]
                         for r in data]
        chart_type   = "pie"
        chart_labels = [r["metodo_pago__descripcion"] or "Sin método" for r in data]
        chart_values = [r["count"] for r in data]
        chart_title  = "Transacciones por método de pago"

    elif tipo == "top_productos":
        data = list(
            DetallePedido.objects
            .filter(pedido__fecha_hora_ingreso__date__gte=desde,
                    pedido__fecha_hora_ingreso__date__lte=hasta)
            .exclude(pedido__estado="cancelado")
            .values("producto__nombre")
            .annotate(cantidad=Sum("cantidad"), ingreso=Sum("subtotal_calculado"))
            .order_by("-cantidad")[:15]
        )
        tabla_headers = ["#", "Producto", "Unidades", "Ingreso ($)"]
        tabla_rows    = [[i, r["producto__nombre"], r["cantidad"], float(r["ingreso"] or 0)]
                         for i, r in enumerate(data, 1)]
        chart_type   = "barh"
        chart_labels = [r["producto__nombre"] for r in data]
        chart_values = [r["cantidad"] for r in data]
        chart_title  = "Productos más vendidos"

    elif tipo == "menos_vendidos":
        from apps.menu.models import Producto as _P
        # Construir mapa {nombre: row} de productos vendidos en el período
        vendidos = {
            r["producto__nombre"]: r
            for r in DetallePedido.objects
            .filter(pedido__fecha_hora_ingreso__date__gte=desde,
                    pedido__fecha_hora_ingreso__date__lte=hasta)
            .exclude(pedido__estado="cancelado")
            .values("producto__nombre")
            .annotate(cantidad=Sum("cantidad"), ingreso=Sum("subtotal_calculado"))
        }
        data = []
        # Iterar sobre TODOS los productos disponibles para incluir los que no se vendieron (cantidad=0)
        for nombre in _P.objects.filter(disponible=True).values_list("nombre", flat=True):
            if nombre in vendidos:
                data.append({"nombre": nombre,
                              "cantidad": vendidos[nombre]["cantidad"],
                              "ingreso": float(vendidos[nombre]["ingreso"] or 0)})
            else:
                data.append({"nombre": nombre, "cantidad": 0, "ingreso": 0.0})
        data.sort(key=lambda x: x["cantidad"])
        data = data[:15]
        tabla_headers = ["#", "Producto", "Unidades", "Ingreso ($)"]
        tabla_rows    = [[i, r["nombre"], r["cantidad"], r["ingreso"]] for i, r in enumerate(data, 1)]
        chart_type   = "barh"
        chart_labels = [r["nombre"] for r in data]
        chart_values = [r["cantidad"] for r in data]
        chart_title  = "Productos menos vendidos"

    elif tipo == "ventas_categoria":
        data = list(
            DetallePedido.objects
            .filter(pedido__fecha_hora_ingreso__date__gte=desde,
                    pedido__fecha_hora_ingreso__date__lte=hasta)
            .exclude(pedido__estado="cancelado")
            .values("producto__categoria__nombre")
            .annotate(total=Sum("subtotal_calculado"), cantidad=Sum("cantidad"))
            .order_by("-total")
        )
        tabla_headers = ["Categoría", "Ventas ($)", "Unidades"]
        tabla_rows    = [[r["producto__categoria__nombre"] or "Sin categoría",
                          float(r["total"] or 0), r["cantidad"]] for r in data]
        chart_type   = "pie"
        chart_labels = [r["producto__categoria__nombre"] or "Sin categoría" for r in data]
        chart_values = [float(r["total"] or 0) for r in data]
        chart_title  = "Ventas por categoría"

    elif tipo == "descuentos_productos":
        data = list(
            DetallePedido.objects
            .filter(pedido__fecha_hora_ingreso__date__gte=desde,
                    pedido__fecha_hora_ingreso__date__lte=hasta)
            .exclude(pedido__estado="cancelado")
            .filter(promocion__isnull=False)
            .values("producto__nombre", "promocion__titulo")
            .annotate(aplicaciones=Count("id"), subtotal=Sum("subtotal_calculado"))
            .order_by("-aplicaciones")[:20]
        )
        tabla_headers = ["Producto", "Promoción", "Aplicaciones", "Subtotal ($)"]
        tabla_rows    = [[r["producto__nombre"], r["promocion__titulo"] or "—",
                          r["aplicaciones"], float(r["subtotal"] or 0)] for r in data]
        chart_type   = "bar"
        chart_labels = [r["producto__nombre"] for r in data]
        chart_values = [r["aplicaciones"] for r in data]
        chart_title  = "Promociones aplicadas"

    elif tipo == "comparativa_empleados":
        meseros = Empleado.objects.filter(rol__in=("mesero", "gerente", "admin"), activo=True)
        data = []
        for emp in meseros:
            agg = base_qs.filter(empleado_entrega=emp).aggregate(
                ventas=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0"), output_field=DecimalField()),
                pedidos=Count("id", distinct=True),
            )
            data.append({"nombre": emp.nombre, "pedidos": agg["pedidos"], "ventas": float(agg["ventas"])})
        data.sort(key=lambda x: x["ventas"], reverse=True)
        tabla_headers = ["Empleado", "Pedidos", "Ventas ($)"]
        tabla_rows    = [[r["nombre"], r["pedidos"], r["ventas"]] for r in data]
        chart_type   = "bar"
        chart_labels = [r["nombre"] for r in data]
        chart_values = [r["ventas"] for r in data]
        chart_title  = "Ventas por empleado"

    elif tipo == "tiempo_atencion":
        from django.db.models import Avg, F, ExpressionWrapper, fields as dj_fields
        pedidos_con_entrega = (
            Pedido.objects
            .filter(fecha_hora_ingreso__date__gte=desde, fecha_hora_ingreso__date__lte=hasta,
                    fecha_hora_entrega__isnull=False)
            .exclude(estado="cancelado")
            .annotate(dur=ExpressionWrapper(
                F("fecha_hora_entrega") - F("fecha_hora_ingreso"),
                output_field=dj_fields.DurationField()
            ))
        )
        global_avg = pedidos_con_entrega.aggregate(avg=Avg("dur"))["avg"]
        global_min = round(global_avg.total_seconds() / 60, 1) if global_avg else 0
        meseros_t = []
        for emp in Empleado.objects.filter(rol__in=("mesero", "gerente", "admin"), activo=True):
            qs  = pedidos_con_entrega.filter(empleado_entrega=emp)
            avg = qs.aggregate(avg=Avg("dur"))["avg"]
            meseros_t.append({
                "nombre":     emp.nombre,
                "pedidos":    qs.count(),
                "tiempo_min": round(avg.total_seconds() / 60, 1) if avg else 0,
            })
        kpis          = [("Tiempo promedio global", f"{global_min} min")]
        tabla_headers = ["Mesero", "Pedidos", "Tiempo promedio (min)"]
        tabla_rows    = [[r["nombre"], r["pedidos"], r["tiempo_min"]] for r in meseros_t]
        chart_type   = "barh"
        chart_labels = [r["nombre"] for r in meseros_t]
        chart_values = [r["tiempo_min"] for r in meseros_t]
        chart_title  = "Tiempo promedio de atención por mesero"

    elif tipo == "mesas_utilizadas":
        data = list(
            Pedido.objects
            .filter(fecha_hora_ingreso__date__gte=desde, fecha_hora_ingreso__date__lte=hasta)
            .exclude(estado="cancelado")
            .values("sesion__mesa__numero_mesa")
            .annotate(
                pedidos=Count("id", distinct=True),
                ventas=Coalesce(Sum("detalles__subtotal_calculado"), Decimal("0"), output_field=DecimalField())
            ).order_by("-pedidos")
        )
        tabla_headers = ["Mesa", "Pedidos", "Ventas ($)"]
        tabla_rows    = [[f"Mesa {r['sesion__mesa__numero_mesa']}", r["pedidos"], float(r["ventas"])]
                         for r in data]
        chart_type   = "bar"
        chart_labels = [f"Mesa {r['sesion__mesa__numero_mesa']}" for r in data]
        chart_values = [r["pedidos"] for r in data]
        chart_title  = "Pedidos por mesa"

    elif tipo == "cancelaciones":
        cancelados = list(
            Pedido.objects
            .filter(fecha_hora_ingreso__date__gte=desde, fecha_hora_ingreso__date__lte=hasta,
                    estado="cancelado")
            .values("id", "fecha_hora_ingreso", "sesion__mesa__numero_mesa",
                    "empleado_entrega__nombre", "motivo_cancelacion")
            .order_by("-fecha_hora_ingreso")
        )
        # Agrupar manualmente por mesero para el gráfico de barras (más flexible que annotate aquí)
        por_mesero: dict = {}
        for c in cancelados:
            nombre = c["empleado_entrega__nombre"] or "Sin asignar"
            por_mesero[nombre] = por_mesero.get(nombre, 0) + 1
        kpis          = [("Total cancelaciones", str(len(cancelados)))]
        tabla_headers = ["ID", "Fecha", "Mesa", "Mesero", "Motivo"]
        tabla_rows    = [
            [
                c["id"],
                str(c["fecha_hora_ingreso"].date()) if c["fecha_hora_ingreso"] else "",
                c["sesion__mesa__numero_mesa"] or "—",
                c["empleado_entrega__nombre"] or "Sin asignar",
                c["motivo_cancelacion"] or "—",
            ]
            for c in cancelados
        ]
        if por_mesero:
            chart_type   = "bar"
            chart_labels = list(por_mesero.keys())
            chart_values = [float(v) for v in por_mesero.values()]
            chart_title  = "Cancelaciones por mesero"

    return {
        "titulo":        titulo,
        "kpis":          kpis,
        "tabla_headers": tabla_headers,
        "tabla_rows":    tabla_rows,
        "tablas_extra":  tablas_extra,
        "chart_type":    chart_type,
        "chart_labels":  chart_labels,
        "chart_values":  chart_values,
        "chart_title":   chart_title,
    }


def _chart_b64(chart_type: str, labels: list, values: list, title: str = "") -> str | None:
    """Genera una gráfica matplotlib y la devuelve como string base64 PNG.

    Usada por exportar_pdf_custom() para incrustar gráficas directamente en el HTML
    del PDF como imágenes base64 (evita dependencias de rutas de archivos en WeasyPrint).

    Parámetros:
        chart_type: "bar", "barh", "line" o "pie".
        labels:     etiquetas del eje X / categorías.
        values:     valores numéricos correspondientes.
        title:      título de la gráfica (opcional).
    Retorna:
        String base64 de la imagen PNG, o None si matplotlib no está disponible
        o si ocurre cualquier error (para no romper la generación del PDF).
    """
    try:
        import matplotlib
        # Backend no interactivo: obligatorio en entornos sin pantalla (servidor/Docker)
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import base64

        # Paleta de verdes corporativa de Mochi Matcha
        PALETTE = ["#2D6A4F", "#52B788", "#95D5B2", "#B7E4C7", "#40916C",
                   "#74C69D", "#1B4332", "#D8F3DC", "#081C15", "#52B788"]

        fig, ax = plt.subplots(figsize=(8, 4))
        # Fondo transparente para integración limpia con el HTML del PDF
        fig.patch.set_alpha(0)
        ax.set_facecolor("none")
        colors = [PALETTE[i % len(PALETTE)] for i in range(len(labels))]

        if chart_type == "bar":
            ax.bar(range(len(labels)), values, color=colors)
            ax.set_xticks(range(len(labels)))
            ax.set_xticklabels(labels, rotation=30, ha="right", fontsize=7)
        elif chart_type == "barh":
            ax.barh(range(len(labels)), values, color=colors)
            ax.set_yticks(range(len(labels)))
            ax.set_yticklabels(labels, fontsize=7)
        elif chart_type == "line":
            ax.plot(range(len(labels)), values, color=PALETTE[0],
                    marker="o", linewidth=2, markersize=5)
            # Área rellena bajo la línea para mayor legibilidad visual
            ax.fill_between(range(len(labels)), values, alpha=0.15, color=PALETTE[0])
            ax.set_xticks(range(len(labels)))
            ax.set_xticklabels(labels, rotation=30, ha="right", fontsize=7)
        elif chart_type == "pie":
            wedges, _, autotexts = ax.pie(
                values, labels=None, colors=colors,
                autopct="%1.0f%%", startangle=90, pctdistance=0.75,
            )
            ax.legend(wedges, labels, loc="center left",
                      bbox_to_anchor=(1, 0.5), fontsize=7)
            for at in autotexts:
                at.set_fontsize(7)

        # Quitar bordes superiores y derechos (estilo minimalista, no aplica a pie)
        if chart_type != "pie":
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_color("#CCCCCC")
            ax.spines["bottom"].set_color("#CCCCCC")
            ax.tick_params(colors="#555555")
            ax.yaxis.grid(True, color="#EEEEEE", linestyle="--", linewidth=0.5)

        if title:
            ax.set_title(title, fontsize=10, color="#2D6A4F", pad=8, fontweight="bold")

        buf = io.BytesIO()
        # dpi=150 equilibra calidad y tamaño del PDF resultante
        fig.savefig(buf, format="png", bbox_inches="tight", dpi=150, transparent=True)
        plt.close(fig)
        buf.seek(0)
        return base64.b64encode(buf.read()).decode()
    except Exception:
        # Si matplotlib no está instalado o falla, se omite la gráfica sin abortar el PDF
        return None


def exportar_excel_custom(desde: date, hasta: date, secciones: list) -> bytes:
    """Genera un Excel con las secciones seleccionadas, incluyendo gráficas openpyxl.

    Crea un libro con una hoja de Resumen más una hoja por cada sección.
    Cada hoja incluye: KPIs, tabla de datos y una gráfica openpyxl nativa.

    Parámetros:
        desde:     fecha de inicio (inclusive).
        hasta:     fecha de fin (inclusive).
        secciones: lista de claves de sección (deben estar en SECCIONES_VALIDAS).
    Retorna:
        bytes del archivo .xlsx listo para servir como descarga.
    Raises:
        ImportError: si openpyxl no está instalado.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    except ImportError:
        raise ImportError("openpyxl no está instalado.")

    VERDE      = "2D6A4F"
    VERDE_PALE = "D8F3DC"

    h_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    h_fill  = PatternFill("solid", fgColor=VERDE)
    alt_fill = PatternFill("solid", fgColor=VERDE_PALE)
    center   = Alignment(horizontal="center", vertical="center")
    left     = Alignment(horizontal="left",   vertical="center")
    thin     = Side(style="thin", color="CCCCCC")
    brd      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(ws, row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font = h_font; c.fill = h_fill; c.alignment = center; c.border = brd
        return c

    def _cell(ws, row, col, value, fmt=None, bold=False, alt=False):
        c = ws.cell(row=row, column=col, value=value)
        c.border = brd; c.alignment = left
        if fmt:   c.number_format = fmt
        if bold:  c.font = Font(bold=True)
        if alt:   c.fill = alt_fill
        return c

    wb = openpyxl.Workbook()

    # ── Hoja Resumen ─────────────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "Resumen"
    ws0.merge_cells("A1:C1")
    t = ws0["A1"]
    t.value = f"Mochi Matcha — Reporte  |  {desde.strftime('%d/%m/%Y')} – {hasta.strftime('%d/%m/%Y')}"
    t.font = Font(name="Calibri", bold=True, size=14, color=VERDE)
    t.alignment = center
    ws0.row_dimensions[1].height = 28

    _hdr(ws0, 3, 1, "SECCIONES INCLUIDAS")
    _hdr(ws0, 3, 2, "PERÍODO")
    for i, sec in enumerate(secciones, start=4):
        alt = (i % 2 == 0)
        _cell(ws0, i, 1, SECCION_TITULOS.get(sec, sec), bold=True, alt=alt)
        if i == 4:
            _cell(ws0, i, 2,
                  f"{desde.strftime('%d/%m/%Y')} – {hasta.strftime('%d/%m/%Y')}", alt=alt)
    ws0.column_dimensions["A"].width = 38
    ws0.column_dimensions["B"].width = 28
    ws0.column_dimensions["C"].width = 18

    # ── Una hoja por sección ─────────────────────────────────────────────────
    for seccion in secciones:
        data  = _datos_seccion(seccion, desde, hasta)
        title = data["titulo"][:31]
        ws    = wb.create_sheet(title)

        n_cols = max(len(data["tabla_headers"]), 2)

        # Título de hoja
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        tc = ws.cell(row=1, column=1, value=data["titulo"])
        tc.font = Font(name="Calibri", bold=True, size=13, color=VERDE)
        tc.alignment = center
        ws.row_dimensions[1].height = 24

        row_ptr = 2

        # KPIs
        if data["kpis"]:
            for lbl, val in data["kpis"]:
                _cell(ws, row_ptr, 1, lbl, bold=True)
                _cell(ws, row_ptr, 2, val)
                row_ptr += 1
            row_ptr += 1

        # Encabezados de tabla principal
        hdr_row = row_ptr
        for j, h in enumerate(data["tabla_headers"], 1):
            _hdr(ws, hdr_row, j, h)
        row_ptr += 1
        data_start = row_ptr

        # Datos de tabla
        for i, row_data in enumerate(data["tabla_rows"]):
            alt = (i % 2 == 1)
            for j, val in enumerate(row_data, 1):
                fmt = "#,##0.00" if isinstance(val, float) and j > 1 else None
                _cell(ws, row_ptr, j, val, fmt=fmt, alt=alt)
            row_ptr += 1

        data_end = row_ptr - 1

        # Anchos de columna
        for j, h in enumerate(data["tabla_headers"], 1):
            max_w = len(str(h))
            for row_data in data["tabla_rows"]:
                if j - 1 < len(row_data):
                    max_w = max(max_w, len(str(row_data[j - 1])))
            ws.column_dimensions[get_column_letter(j)].width = min(max_w + 4, 55)

        # Tablas extra (corte_dia)
        if data.get("tablas_extra"):
            row_ptr += 2
            for extra in data["tablas_extra"]:
                ws.merge_cells(start_row=row_ptr, start_column=1,
                               end_row=row_ptr, end_column=len(extra["headers"]))
                ec = ws.cell(row=row_ptr, column=1, value=extra["titulo"])
                ec.font = Font(bold=True, color=VERDE, size=11)
                row_ptr += 1
                for j, h in enumerate(extra["headers"], 1):
                    _hdr(ws, row_ptr, j, h)
                row_ptr += 1
                for i, row_data in enumerate(extra["rows"]):
                    alt = (i % 2 == 1)
                    for j, val in enumerate(row_data, 1):
                        fmt = "#,##0.00" if isinstance(val, float) and j > 1 else None
                        _cell(ws, row_ptr, j, val, fmt=fmt, alt=alt)
                    row_ptr += 1
                row_ptr += 1

        # ── Gráfica openpyxl ──────────────────────────────────────────────
        if data["chart_type"] and data["chart_labels"] and data["chart_values"]:
            n_chart = len(data["chart_labels"])
            # Columna auxiliar fuera del rango visible para almacenar datos de la gráfica
            helper_col = n_cols + 2

            # Escribir datos auxiliares para la gráfica (openpyxl requiere referencias a celdas)
            ws.cell(row=hdr_row,   column=helper_col - 1, value="Etiqueta")
            ws.cell(row=hdr_row,   column=helper_col,     value="Valor")
            for k, (lbl, val) in enumerate(zip(data["chart_labels"], data["chart_values"]), 1):
                ws.cell(row=hdr_row + k, column=helper_col - 1, value=lbl)
                ws.cell(row=hdr_row + k, column=helper_col,     value=val)

            # Referencias al rango de datos y categorías para asociar a la gráfica
            val_ref  = Reference(ws, min_col=helper_col, max_col=helper_col,
                                  min_row=hdr_row, max_row=hdr_row + n_chart)
            cat_ref  = Reference(ws, min_col=helper_col - 1,
                                  min_row=hdr_row + 1, max_row=hdr_row + n_chart)

            if data["chart_type"] in ("bar", "barh"):
                chart = BarChart()
                # "col" = barras verticales, "bar" = horizontales (terminología openpyxl)
                chart.type = "col" if data["chart_type"] == "bar" else "bar"
                chart.grouping = "clustered"
                chart.style    = 10
                chart.add_data(val_ref, titles_from_data=True)
                chart.set_categories(cat_ref)
            elif data["chart_type"] == "line":
                chart = LineChart()
                chart.style = 10
                chart.add_data(val_ref, titles_from_data=True)
                chart.set_categories(cat_ref)
            else:  # pie
                chart = PieChart()
                chart.style = 10
                chart.add_data(val_ref, titles_from_data=True)
                chart.set_categories(cat_ref)

            chart.title  = data["chart_title"]
            chart.width  = 18
            chart.height = 12
            # Anclar la gráfica dos columnas a la derecha de los datos auxiliares
            anchor = get_column_letter(helper_col + 2) + str(hdr_row)
            ws.add_chart(chart, anchor)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_pdf_custom(desde: date, hasta: date, secciones: list) -> bytes:
    """Genera un PDF con las secciones seleccionadas, con gráficas matplotlib.

    Para cada sección: obtiene datos con _datos_seccion(), genera la gráfica
    como imagen base64 con _chart_b64() y renderiza el template HTML que
    WeasyPrint convierte a PDF.

    Parámetros:
        desde:     fecha de inicio (inclusive).
        hasta:     fecha de fin (inclusive).
        secciones: lista de claves de sección (deben estar en SECCIONES_VALIDAS).
    Retorna:
        bytes del archivo PDF listo para servir como descarga.
    Raises:
        ImportError: si weasyprint no está instalado.
        RuntimeError: si la renderización HTML→PDF falla.
    """
    try:
        import weasyprint
    except ImportError:
        raise ImportError("weasyprint no está instalado.")

    secciones_data = []
    for seccion in secciones:
        data       = _datos_seccion(seccion, desde, hasta)
        grafica_b64 = None
        # Generar gráfica solo si hay datos de chart; si matplotlib falla devuelve None
        if data["chart_type"] and data["chart_labels"] and data["chart_values"]:
            grafica_b64 = _chart_b64(
                data["chart_type"],
                data["chart_labels"],
                data["chart_values"],
                data["chart_title"],
            )
        secciones_data.append({
            "titulo":        data["titulo"],
            "kpis":          data["kpis"],
            "tabla_headers": data["tabla_headers"],
            "tabla_rows":    data["tabla_rows"],
            "tablas_extra":  data.get("tablas_extra", []),
            "grafica_b64":   grafica_b64,
        })

    # Renderizar el template HTML con todos los datos de las secciones
    html_str = render_to_string("gerente/reporte_pdf_custom.html", {
        "desde":          desde,
        "hasta":          hasta,
        "secciones_data": secciones_data,
    })
    try:
        return weasyprint.HTML(string=html_str).write_pdf()
    except Exception as exc:
        raise RuntimeError(f"Error al generar PDF: {exc}") from exc
